import os
import re
import openpyxl
class inner_functions_class():
    def __init__(self):
        pass
    def get_the_word_inbetween(self,text, start_char, end_char):
        start_index = text.find(start_char)
        if start_index == -1:
            return None
        end_index = text.find(end_char, start_index + 1)
        if end_index == -1 or end_index <= start_index:
            return None
        return text[start_index + 1:end_index]
    def count_occurrences(self,word:str, string:str):
        count = 0
        word_len = len(word)
        text_len = len(string)
        for i in range(text_len - word_len + 1):
            if string[i:i + word_len] == word:
                count += 1
        return count
    def get_line_of_phrase_in_text(self,text, phrase):
        lines = text.splitlines()
        for line in lines:
            if phrase in line:
                line_without_phrase = line.replace(phrase, "")
                return line_without_phrase.strip()
        return None
    def modify_line_containing_word(self,text, word, new_line_content):
        lines = text.splitlines()
        line_number = -1
        for i, line in enumerate(lines):
            if word in line:
                line_number = i
        if line_number != -1:
            lines[line_number] = new_line_content
            return "\n".join(lines)
        else:
            return text
    def group_by_element(self,input_list,index):
        grouped_list = {}
        for sublist in input_list:
            second_element = sublist[index]
            if second_element not in grouped_list:
                grouped_list[second_element] = []
            grouped_list[second_element].append(sublist)
        return list(grouped_list.values())
    def add_data_to_inner_lists(self,main_list,second_list):
        result = []
        second_index = 0
        for inner_item in main_list:
            if second_index < len(second_list):
                result.append(inner_item + [second_list[second_index]])
                second_index += 1
            else:
                result.append(inner_item + [None])
                print("Warning: second_list is shorter than expected. Filling with None.")
        return result
    def combine_lists(self,input_list):
        output_list = []
        for inner_list in input_list:
            for item in inner_list:
                output_list.append(item)
        return output_list
inner_functions = inner_functions_class()
class create_class():
    def __init__(self):
        pass
    def makeDB(self,newfile:str):
        if newfile[-4:] == '.pdb' or newfile[-4:] == '.PDB':
            makeDBX = open(newfile,'x')
        else:
            makeDBX = open(f'{newfile}.pdb', 'x')
        makeDBX.write('#POWER_DB')
        makeDBX.close()
    def makecontainer(self,file:str,name:str):
        scancontainers = open(file,'r')
        r = scancontainers.read()
        scancontainers.close()
        num = inner_functions.count_occurrences('$<', r)
        if f',{name}>' not in r:
            makecontainer = open(file, 'a')
            if num == 0:
                if r.endswith('\n'):
                   makecontainer.write(f"$<0,{name}>")
                else:
                   makecontainer.write(f"\n$<0,{name}>")
            else:
                if r.endswith('\n'):
                   makecontainer.write(f"$<{num},{name}>")
                else:
                   makecontainer.write(f"\n$<{num},{name}>")
            makecontainer.close()
        else:
            pass
    def maketable(self,file:str,name:str):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        num = inner_functions.count_occurrences('&<', r)
        if f'^{name}>' not in r:
            makecontainer = open(file, 'a')
            if num == 0:
                if r.endswith('\n'):
                   makecontainer.write(f"&<0^{name}>")
                else:
                    makecontainer.write(f"\n&<0^{name}>")
            else:
                if r.endswith('\n'):
                   makecontainer.write(f"&<{num}^{name}>")
                else:
                    makecontainer.write(f"\n&<{num}^{name}>")
            makecontainer.close()
        else:
            pass
create = create_class()
class container_data_class():
    def __init__(self):
        pass
    def getname(self, file: str, id:int):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if f'$<{id},' in r:
            return inner_functions.get_the_word_inbetween(f'$<{id},'+inner_functions.get_line_of_phrase_in_text(r,f'$<{id},'), ',', '>')
    def getid(self,file:str,name:str,plogic:bool=True):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        data = ''
        i = 0
        while True:
            if f'$<{i},{name}>' in r:
                data = f'$<{i},{name}>'
                i = i + 1
                break
            else:
                break
        if data != '':
            return int(inner_functions.get_the_word_inbetween(data, '<', ',')) if plogic else int(
                inner_functions.get_the_word_inbetween(data, '<', ',')) + 1
        else:
            return -1 if plogic else 0
    def insert(self,file:str,data:str,address=None,showrelational:bool=False):
        if address is None:
            address = []
        containerid = address[0]
        sectorid = address[1]
        info = self.numbersectors(file,0)
        if showrelational:
            print(sectorid,info)
        makecontainer = open(file, 'a')
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if not other.check(file,'sector',[containerid,sectorid]):
            if sectorid - info <= 1:
               if r.endswith('\n'):
                  makecontainer.write(f"!<[{containerid},{sectorid}],{data}>!")
               else:
                  makecontainer.write(f"\n!<[{containerid},{sectorid}],{data}>!")
        else:
            pass
        makecontainer.close()
    def read(self,file:str,address=None):
        if address is None:
            address = []
        containerid = address[0]
        sectorid = address[1]
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        data = ""
        if f'!<[{containerid},{sectorid}]' in r:
            data = inner_functions.get_line_of_phrase_in_text(r,f'!<[{containerid},{sectorid}]')[1:-2]
        return data
    def edit(self,file:str,data:str,address=None):
        if address is None:
            address = []
        containerid = address[0]
        sectorid = address[1]
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if other.check(file, 'sector', [containerid, sectorid]):
            actdata = inner_functions.modify_line_containing_word(r,f'!<[{containerid},{sectorid}]',f'!<[{containerid},{sectorid}],{data}>!')
            rccontainers = open(file, 'w')
            rccontainers.write('')
            rccontainers.close()
            editcontainers = open(file, 'w')
            editcontainers.write(actdata)
            editcontainers.close()
        else:
            pass
    def change_name(self,file:str,new_name:str,containerid):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if other.check(file, 'container', [containerid, self.getname(file,containerid)]):
            actdata = inner_functions.modify_line_containing_word(r,f'$<{containerid},{self.getname(file,containerid)}>',f'$<{containerid},{new_name}>')
            rccontainers = open(file, 'w')
            rccontainers.write('')
            rccontainers.close()
            editcontainers = open(file, 'w')
            lines = actdata.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            actdatan = '\n'.join(non_empty_lines)
            editcontainers.write(actdatan)
            editcontainers.close()
        else:
            pass
    def readsectors(self,file:str,containerid:int):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        data = []
        i = 0
        while True:
            if f'!<[{containerid},{i}]' in r:
                data.append(inner_functions.get_line_of_phrase_in_text(r, f'!<[{containerid},{i}]')[1:-2])
                i = i + 1
            else:
                break
        return data
    def numbercontainers(self, file: str,plogic:bool=False):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if plogic is False:
            return inner_functions.count_occurrences('$<', r)
        else:
            return inner_functions.count_occurrences('$<', r)-1
    def numbersectors(self, file: str,containerid:int,plogic:bool=False):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if plogic is False:
            return inner_functions.count_occurrences(f'!<[{containerid}', r)
        else:
            return inner_functions.count_occurrences(f'!<[{containerid}', r)-1
    def delete(self,file:str,address=None):
        if address is None:
            address = []
        containerid = address[0]
        sectorid = address[1]
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if other.check(file, 'sector', [containerid, sectorid]):
            actdata = inner_functions.modify_line_containing_word(r,f'!<[{containerid},{sectorid}]',f'')
            rccontainers = open(file, 'w')
            rccontainers.write('')
            rccontainers.close()
            editcontainers = open(file, 'w')
            lines = actdata.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            actdatan = '\n'.join(non_empty_lines)
            editcontainers.write(actdatan)
            editcontainers.close()
        else:
            pass
    def drop(self,file:str,containerid:int):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        endata = ''
        secnum = container_data.numbersectors(file,containerid)
        atdata = inner_functions.modify_line_containing_word(r, f'$<{containerid},', f'')
        ik = 0
        if secnum != -1:
            cha = ''
            while True:
                if f'!<[{containerid},{ik}]' in atdata:
                    actdata = inner_functions.modify_line_containing_word(atdata, f'!<[{containerid},{ik}]', f'')
                    atdata = actdata
                    cha = actdata
                else:
                    if cha == '':
                        actdata = atdata
                    else:
                        actdata = cha
                    break
                if ik == secnum:
                    endata = actdata
                    break
                ik = ik + 1
        else:
            endata = atdata
        rccontainers = open(file, 'w')
        rccontainers.write('')
        rccontainers.close()
        editcontainers = open(file, 'w')
        lines = endata.split('\n')
        non_empty_lines = [line for line in lines if line.strip() != '']
        actdatan = '\n'.join(non_empty_lines)
        editcontainers.write(actdatan)
        editcontainers.close()
container_data = container_data_class()
class table_data_class():
    def __init__(self):
        pass
    def getname(self, file: str, id:int):
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        if f'&<{id}^' in r:
            return inner_functions.get_the_word_inbetween(f'&<{id}^'+inner_functions.get_line_of_phrase_in_text(r,f'&<{id}^'), '^', '>')
    def getid(self, file: str, name: str, plogic: bool = True):
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        data = ''
        i = 0
        while True:
            if f'&<{i}^{name}>' in r:
                data = f'&<{i}^{name}>'
                i = i + 1
                break
            else:
                break
        if data != '':
            return int(inner_functions.get_the_word_inbetween(data, '<', '^')) if plogic else int(
                inner_functions.get_the_word_inbetween(data, '<', '^')) + 1
        else:
            return -1 if plogic else 0

    def hcolumn(self, file: str, tableid: int, plogic: bool = False, sprow: int = -1):
        scantables = open(file, 'r')
        lines = scantables.read().strip().split('\n')
        scantables.close()

        matches = []

        if sprow == -1:
            for line in lines:
                pattern = rf'~<\[{tableid};(\d+)\?.*]'
                match = re.search(pattern, line)
                if match:
                    matches.append(match.group(1))
        else:
            for line in lines:
                pattern = rf'~<\[{tableid};(\d+)\?{sprow}\]'
                match = re.search(pattern, line)
                if match:
                    matches.append(match.group(1))

        if not matches:
            if plogic:
                return -1
            else:
                return 0

        try:
            max_col = max(map(int, matches))
            if plogic:
                return max_col
            else:
                return max_col + 1
        except ValueError:
            if plogic:
                return -1
            else:
                return 0
    def hrow(self, file: str, tableid: int, plogic: bool = False, sprow: int = -1):
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()

        lines = r.strip().split('\n')
        matches = []

        if sprow == -1:
            for line in lines:
                pattern = rf'~<\[{tableid};\d+\?(\d+)]'
                match = re.search(pattern, line)
                if match:
                    matches.append(match.group(1))
        else:
            for line in lines:
                pattern = rf'~<\[{tableid};{sprow}\?(\d+)]'
                match = re.search(pattern, line)
                if match:
                    matches.append(match.group(1))

        if not matches:
            if plogic:
                return
            else:
                return 0

        try:
            max_row = max(map(int, matches))
            if plogic:
                return max_row
            else:
                return max_row + 1
        except ValueError:
            if plogic:
                return -1
            else:
                return 0
    def numbertables(self,file:str,plogic:bool=True):
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        if plogic is False:
            return inner_functions.count_occurrences('&<',r)
        else:
            return inner_functions.count_occurrences('&<', r)-1
    def numbercolumns(self,file:str,address=None,plogic:bool=False):
        return self.hcolumn(file, address[0], plogic, address[1])
    def numberrows(self,file:str,address=None,plogic:bool=False):
        return self.hrow(file, address[0], plogic, address[1])
    def totalcolumns(self,file:str,tableid:int,plogic:bool=False):
        return self.hcolumn(file, tableid, plogic)
    def totalrows(self,file:str,tableid:int,plogic:bool=False):
        return self.hrow(file, tableid, plogic)
    def totaltable(self,file:str,tableid:int,plogic:bool=False):
        return [self.hcolumn(file, tableid, plogic), self.hrow(file, tableid, plogic)]
    def insert(self,file:str,data:str,address=None,showmatrix:bool=False):
        if address is None:
            address = []
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        maketable = open(file, 'a')
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        if r.endswith('\n\n'): r = r[:-1]
        info = self.totaltable(file, tableid)
        if showmatrix:
            print(columnid,info[0])
            print(rowid,info[1])
        if not other.check(file,'cell',[tableid,columnid,rowid]):
            if columnid - info[0] <= 1:
               if rowid - info[1] <= 1:
                  if r.endswith('\n'):
                     maketable.write(f"~<[{tableid};{columnid}?{rowid}],{data}>~")
                  else:
                     maketable.write(f"\n~<[{tableid};{columnid}?{rowid}],{data}>~")
        else:
            pass
        maketable.close()
    def read(self,file:str,address=None):
        if address is None:
            address = []
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        data = ""
        if other.check(file, 'cell', [tableid, columnid, rowid]):
            if f'~<[{tableid};{columnid}?{rowid}]' in r:
                data = inner_functions.get_line_of_phrase_in_text(r,f'~<[{tableid};{columnid}?{rowid}]')[1:-2]
            return data
        else:
            pass
    def readcolumns(self,file:str,address=None):
        tableid = address[0]
        rowid = address[1]
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        data = []
        i = 0
        while True:
            if f'~<[{tableid};{i}?{rowid}]' in r:
                data.append(inner_functions.get_line_of_phrase_in_text(r, f'~<[{tableid};{i}?{rowid}]')[1:-2])
                i = i + 1
            else:
                break
        return data
    def readrows(self,file:str,address=None):
        tableid = address[0]
        columnid = address[1]
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        data = []
        i = 0
        while True:
            if f'~<[{tableid};{columnid}?{i}]' in r:
                data.append(inner_functions.get_line_of_phrase_in_text(r, f'~<[{tableid};{columnid}?{i}]')[1:-2])
                i = i + 1
            else:
                break
        return data
    def edit(self,file:str,data:str,address=None):
        if address is None:
            address = []
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        if other.check(file, 'cell', [tableid, columnid, rowid]):
            actdata = inner_functions.modify_line_containing_word(r,f'~<[{tableid};{columnid}?{rowid}]',f'~<[{tableid};{columnid}?{rowid}],{data}>')
            rctables = open(file, 'w')
            rctables.write('')
            rctables.close()
            edittables = open(file, 'w')
            lines = actdata.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            actdatan = '\n'.join(non_empty_lines)
            edittables.write(actdatan)
            edittables.close()
        else:
            pass
    def change_name(self,file:str,new_name:str,tableid):
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        if other.check(file, 'table', [tableid, self.getname(file,tableid)]):
            actdata = inner_functions.modify_line_containing_word(r,f'&<{tableid}^{self.getname(file,tableid)}>',f'&<{tableid}^{new_name}>')
            rctables = open(file, 'w')
            rctables.write('')
            rctables.close()
            edittables = open(file, 'w')
            lines = actdata.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            actdatan = '\n'.join(non_empty_lines)
            edittables.write(actdatan)
            edittables.close()
        else:
            pass
    def all_addresses_grouping(self,file:str,tableid:int,filtermode:int):
        #filtermode must be either 0(columns) or 1(rows)
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        raw = []
        data1 = []
        info = self.totaltable(file,tableid,False)
        for i in range(info[0]):
            if f'~<[{tableid};{i}' in r:
                for iu in range(info[1]):
                    if f'~<[{tableid};{i}?{iu}]' in r:
                        raw.append([tableid,i,iu])
                    else:
                        break
            else:
                break
        if filtermode < 2:
            return inner_functions.group_by_element(raw,filtermode+1)
        else:
            return []
    def all_addresses_list(self,file:str,tableid:int,totalnum:bool=False):
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        raw = []
        info = self.totaltable(file,tableid,False)
        for i in range(info[0]):
            if f'~<[{tableid};{i}' in r:
                for iu in range(info[1]):
                    if f'~<[{tableid};{i}?{iu}]' in r:
                        raw.append([tableid,i,iu])
                    else:
                        break
            else:
                break
        if totalnum:
            return len(raw)
        else:
            return raw
    def delete(self,file:str,address=None):
        if address is None:
            address = []
        tableid = address[0]
        columnid = address[1]
        rowid = address[2]
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        if other.check(file, 'cell', [tableid, columnid, rowid]):
            actdata = inner_functions.modify_line_containing_word(r,f'~<[{tableid};{columnid}?{rowid}]',f'')
            rctables = open(file, 'w')
            rctables.write('')
            rctables.close()
            edittables = open(file, 'w')
            lines = actdata.split('\n')
            non_empty_lines = [line for line in lines if line.strip() != '']
            actdatan = '\n'.join(non_empty_lines)
            edittables.write(actdatan)
            edittables.close()
        else:
            pass
    def drop(self,file:str,tableid:int):
        scantables = open(file, 'r')
        r = scantables.read()
        scantables.close()
        endata = ''
        secnum = self.hrow(file,tableid)
        atdata = inner_functions.modify_line_containing_word(r, f'&<{tableid}^', f'')
        ik = 0
        if secnum != -1:
            while True:
               c = 0
               cha = ''
               while True:
                  if f'~<[{tableid};{c}?{ik}]' in atdata:
                     actdata = inner_functions.modify_line_containing_word(atdata,f'~<[{tableid};{c}?{ik}]',f'')
                     atdata = actdata
                     cha = actdata
                  else:
                      if cha == '':
                         actdata = atdata
                      else:
                         actdata = cha
                      break
                  c = c + 1
               if ik == secnum:
                  endata = actdata
                  break
               ik = ik + 1
        else:
            endata = atdata
        rctables = open(file, 'w')
        rctables.write('')
        rctables.close()
        edittables = open(file, 'w')
        lines = endata.split('\n')
        non_empty_lines = [line for line in lines if line.strip() != '']
        actdatan = '\n'.join(non_empty_lines)
        edittables.write(actdatan)
        edittables.close()

    def export_tables_to_excel(self, dbfile:str, filepath:str):
        num_sheets = self.numbertables(dbfile, False)
        total_items = 0
        for table_id in range(num_sheets):
            total_items += self.all_addresses_list(dbfile, table_id, True)
        ata_list = []
        raw_data = []
        stuff_list = []
        for main in range(num_sheets):
            raw_data.append(self.all_addresses_list(dbfile,main))
        sraw_data = inner_functions.combine_lists(raw_data)
        for m in range(total_items):
            stuff_list.append(self.read(dbfile, [sraw_data[m][0], sraw_data[m][1], sraw_data[m][2]]))
        ata_list.append(inner_functions.add_data_to_inner_lists(sraw_data,stuff_list))
        data_list = inner_functions.combine_lists(ata_list)
        try:
            created_sheets = {}

            if os.path.isfile(filepath):
                workbook = openpyxl.load_workbook(filepath)
            else:
                workbook = openpyxl.Workbook()
                workbook = openpyxl.Workbook()
                std = workbook['Sheet']
                workbook.remove(std)
            for item in range(len(data_list)):
                if len(data_list[item]) == 4:
                    table_id, col_id_0, row_id_0, value = data_list[item]
                    table_name = self.getname(dbfile, table_id)
                    if table_name not in created_sheets:
                        if table_name not in workbook.sheetnames:
                            workbook.create_sheet(table_name)
                        created_sheets[table_name] = True

                    if table_name in workbook.sheetnames:
                        sheet = workbook[table_name]
                        col_id_1 = col_id_0 + 1
                        row_id_1 = row_id_0 + 1
                        sheet.cell(row=row_id_1, column=col_id_1, value=value)
                    else:
                        print(f"Warning: Sheet '{table_name}' not found.")
                else:
                    print(f"Warning: Invalid data item {item}. Expected [table_id, column_id, row_id, value].")
            workbook.save(filepath)
            print(f"Data inserted into '{filepath}'.")
        except Exception as e:
            print(f"An error occurred: {e}")
table_data = table_data_class()
class other_class():
    def __init__(self):
        pass
    def clear(self,file:str):
        rccontainers = open(file, 'w')
        rccontainers.write('')
        rccontainers.close()
        accontainers = open(file, 'w')
        accontainers.write('#POWER_DB')
        accontainers.close()
    def check(self, file:str, itemtype:str, address=None):
        if address is None:
            address = []
        scancontainers = open(file, 'r')
        r = scancontainers.read()
        scancontainers.close()
        if itemtype.lower() == 'container':
            containerid = address[0]
            name = address[1]
            if f'$<{containerid},{name}>' in r:
                return True
            else:
                return False
        if itemtype.lower() == 'table':
            tableid = address[0]
            name = address[1]
            if f'&<{tableid}^{name}>' in r:
                return True
            else:
                return False
        if itemtype.lower() == 'sector':
            containerid = address[0]
            sectorid = address[1]
            if f'!<[{containerid},{sectorid}],' in r:
                return True
            else:
                return False
        if itemtype.lower() == 'cell':
            tableid = address[0]
            columnid = address[1]
            rowid = address[2]
            if f'~<[{tableid};{columnid}?{rowid}],' in r:
                return True
            else:
                return False
other = other_class()