import re
from pypinyin import lazy_pinyin
import pykakasi
import numpy as np
import wavio
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from dvatioff_audio.voice.azure_speech2text import speech_to_text
from dvatioff_audio.utils import get_desktop_path, text_wrong_proportion, text_distance


def generate_sine_wave(frequency=440.0, duration=1, rate=8000):
    """生成一个正弦波形"""
    t = np.linspace(0, duration, int(duration * rate), endpoint=False)
    return np.sin(2 * np.pi * frequency * t)


def save_sine_wave_to_wav(filename, frequency=440.0, duration=1, rate=8000):
    """保存正弦波到.wav文件"""
    sine_wave = generate_sine_wave(frequency, duration, rate)
    wavio.write(filename, sine_wave, rate, sampwidth=3)


def chinese_to_pinyin(text):
    """
    把中文转化为拼音
    """
    lazy_pinyin_text = lazy_pinyin(text)
    sentence = " ".join(lazy_pinyin_text)

    # 将连续的空白字符替换为单个空格
    sentence = re.sub(r'\s+', ' ', sentence)

    return sentence


def japanese_to_kata(text):
    """
    将日文转化为片假名
    """
    kks = pykakasi.kakasi()
    kata_text = kks.convert(text)
    kata = ''

    for item in kata_text:
        kata += item['kana']

    return kata


def japanese_to_romaji(text):
    """
    将日文转化为罗马音
    """
    kks = pykakasi.kakasi()
    kks_text = kks.convert(text)
    romaji = ''

    for item in kks_text:
        romaji += item['hepburn']

    return romaji


def remove_text_within_brackets(text):
    """
    去除文本中括号内的内容，包括括号本身，这里的括号包括半角括号 () 和全角括号 （）和方括号 []
    """
    pattern = re.compile(r'\(.*?\)|（.*?）|\[.*?\]')

    return pattern.sub('', text)


def replace_punctuation_to_space(text):
    """
    去除文本中所有的标点符号
    """
    text = text.replace('\n', '')

    # 正则表达式匹配所有标点符号
    pattern = re.compile(r'[!"#$%&\'()*+,-./:;<=>?@\[\\\]^_`{|}~，？！…。、【】；‘’“”·（）—―～「」]')

    text = pattern.sub(' ', text)
    text = re.sub(r'\s+', ' ', text)

    return text.strip()


def replace_punctuation_to_empty(text):
    """
    去除文本中的标点符号
    """

    # 将'\n'替换为''
    text = text.replace('\n', '')

    # 正则表达式匹配所有标点符号
    pattern = re.compile(r'[!"#$%&\'()*+,-./:;<=>?@\[\\\]^_`{|}~，？！…。、【】；‘’“”·（）—―「」]')
    text = pattern.sub('', text)

    return text


def remove_all_space(text):
    """
    去除文本中的所有空格
    """
    text = text.replace(' ', '')

    return text


def text_speech_verify_worker(folder_path, legal_audio_stt_dict, language, subscription, region, update_queue):
    """
    遍历文件夹中的所有 wav 文件，使用 Azure Speech2Text 接口识别音频中的文本，并与合法的文本进行比对
    """

    total_index = 0
    total_file_num = len(legal_audio_stt_dict)

    for root, dirs, files in os.walk(folder_path):
        cur_index = 0
        cur_folder_file_num = 0

        # 统计当前子文件夹中的合法 wav 文件数量
        for file in files:
            if file.endswith('.wav'):
                file_name, file_extension = os.path.splitext(file)
                if file_name in legal_audio_stt_dict.keys():
                    cur_folder_file_num += 1

        for index, file in enumerate(files, start=1):
            result = {}
            if file.endswith('.wav'):
                cur_index += 1
                total_index += 1
                file_name, file_extension = os.path.splitext(file)
                if file_name not in legal_audio_stt_dict.keys():
                    continue
                file_path = os.path.join(root, file)
                # 获取父目录的完整路径
                parent_directory_path = os.path.dirname(file_path)
                # 获取父目录的名称
                parent_directory_name = os.path.basename(parent_directory_path)

                result["parent_directory_name"] = parent_directory_name
                result["file_name"] = file_name
                result["file_path"] = file_path
                result["total_index"] = total_index
                result["total_file_num"] = total_file_num
                result["cur_index"] = cur_index
                result["cur_folder_file_num"] = cur_folder_file_num

                if "Shout" in file_name:
                    true_text = "本句为呼喝声，跳过检测"
                    result["true_text"] = true_text
                    update_queue.put(result)
                    continue

                # 识别音频中的文本
                speech_text = speech_to_text(file_path, language, subscription, region)
                if speech_text is None:
                    true_text = "Azure 接口调用失败"
                    result["true_text"] = true_text
                    update_queue.put(result)
                    continue

                # 去除 stt文本/合法文本 中所有的标点符号，如果语言为英语还须将文本转化为小写
                speech_text = remove_text_within_brackets(speech_text)
                speech_text = replace_punctuation_to_empty(speech_text)
                true_text = remove_text_within_brackets(legal_audio_stt_dict[file_name]["dialogue"])
                true_text = replace_punctuation_to_empty(true_text)
                speech_text = speech_text.lower() if language == "English" else remove_all_space(speech_text)
                true_text = true_text.lower() if language == "English" else remove_all_space(true_text)

                result["true_text"] = true_text
                result["speech_text"] = speech_text

                if language == "Chinese":
                    # 将中文文本转化为拼音
                    true_pinyin = chinese_to_pinyin(true_text)
                    speech_pinyin = chinese_to_pinyin(speech_text)

                    # 计算两个文本之间的词错误率
                    wer = text_wrong_proportion(true_pinyin, speech_pinyin)
                    wer = 1 if wer >= 1 else wer

                    result["true_pinyin"] = true_pinyin
                    result["speech_pinyin"] = speech_pinyin
                    result["wer"] = wer
                elif language == "Japanese":
                    # 将日文文本转化为片假名和罗马音
                    true_kata = japanese_to_kata(true_text)
                    file_kata = japanese_to_kata(speech_text)
                    true_kata_romaji = japanese_to_romaji(true_kata)
                    file_kata_romaji = japanese_to_romaji(file_kata)

                    # 计算两个文本之间的词错误率
                    wer_text = text_wrong_proportion(true_text, speech_text)
                    wer_kata = text_wrong_proportion(true_kata, file_kata)
                    distance = text_distance(true_kata_romaji, file_kata_romaji)
                    distance = distance if distance > 0 else 0

                    # 计算两个文本之间的 Levenshtein 距离
                    text_spaceless = true_kata_romaji.replace(' ', '')
                    text_length = len(text_spaceless)

                    # 计算最终的词错误率
                    wer_romaji = distance / text_length if text_length > 0 else 1
                    wer_romaji = wer_romaji if wer_romaji >= 0 else 1
                    wer = min(wer_text, wer_kata, wer_romaji)
                    wer = 1 if wer >= 1 else wer
                    result["true_kata"] = true_kata
                    result["speech_kata"] = file_kata
                    result["true_kata_romaji"] = true_kata_romaji
                    result["speech_kata_romaji"] = file_kata_romaji
                    result["wer"] = wer
                elif language == "Korean":
                    # 去除文本中所有的空格
                    distance = text_distance(remove_all_space(true_text), remove_all_space(speech_text))

                    # 计算两个文本之间的 Levenshtein 距离
                    distance = distance if distance > 0 else 0
                    text_spaceless = remove_all_space(true_text)
                    text_length = len(text_spaceless)

                    # 计算两个文本之间的词错误率
                    wer = distance / text_length if text_length > 0 else 1
                    wer = 1 if wer >= 1 else wer
                    result["wer"] = wer
                elif language == "Russian":
                    # 计算两个文本之间的词错误率
                    distance = text_distance(true_text, speech_text)
                    text_spaceless = true_text.replace(' ', '')
                    text_length = len(text_spaceless)

                    # 计算两个文本之间的词错误率
                    wer = distance / text_length if text_length > 0 else 1
                    wer = 1 if wer >= 1 else wer
                    result["wer"] = wer
                else:
                    # 计算两个文本之间的词错误率
                    wer = text_wrong_proportion(true_text, speech_text)
                    wer = 1 if wer >= 1 else wer
                    result["wer"] = wer

                update_queue.put(result)


def output_text_speech_verify_result(folder_name, file_list, error_log, language):
    """
    整合 Speech2Text 的结果中匹配率异常的 wav 文件，并生成 xlsx 错误日志，发送至桌面
    """

    if language == "Chinese":
        hyperlink_column = 7
    elif language == "Japanese":
        hyperlink_column = 9
    else:
        hyperlink_column = 5

    new_paths = []
    # 获取桌面路径
    desktop_path = get_desktop_path()

    desktop_folder_name = f'{language}-{folder_name}-错误文件'
    desktop_folder_path = os.path.join(desktop_path, desktop_folder_name)

    # 在桌面创建新文件夹
    if not os.path.exists(desktop_folder_path):
        os.makedirs(desktop_folder_path)

    log_name = '错误日志.xlsx'
    log_path = os.path.join(desktop_folder_path, log_name)

    for file in file_list:
        # 仅当文件存在时执行复制操作
        if os.path.isfile(file):
            shutil.copy(file, desktop_folder_path)
            new_paths.append(os.path.basename(file))

    # 将原始数据分割为单独的记录
    records = error_log.strip().split('\n\n')

    # 解析每个记录
    parsed_data = []
    for record in records:
        lines = record.split('\n')
        record_dict = {}
        for line in lines:
            line = line.replace('|', '\n')
            key, value = line.split(': ')
            record_dict[key] = value
        parsed_data.append(record_dict)

    # 创建DataFrame
    df = pd.DataFrame(parsed_data)

    # 导出到Excel
    df.to_excel(log_path, index=False)

    # 使用openpyxl打开Excel文件
    wb = load_workbook(log_path)
    ws = wb.active

    # 为每个文件路径设置超链接
    for row, file_path in enumerate(new_paths, start=2):  # start=1 从第一行开始
        cell = ws.cell(row=row, column=hyperlink_column)  # 第5列
        cell.value = "Open File"
        cell.hyperlink = file_path
        cell.font = Font(color='0000FF', underline='single')

    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 16

    # 保存更改
    wb.save(log_path)
