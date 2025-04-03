from typing import Dict, List, Optional, Set, Union 
from openpyxl import Workbook as _Workbook
from openpyxl.styles import (
    Alignment as _Alignment,
    Border as _Border,
    Font as _Font,
    NamedStyle as _NamedStyle,
    PatternFill as _PatternFill,
    Side as _Side,
)
from openpyxl.styles.colors import Color as _Color
from openpyxl.utils import get_column_letter

from .errors import CellmateException


class Style:
    """
    Defines a reusable style for Excel cells, encapsulating font, border,
    fill, alignment, and number format properties.

    Parameters
    ----------
    name : str
        The name of the style.
    font : dict, optional
        Dictionary containing font attributes such as 'name', 'bold', 'italic',
        'size', and 'color' (default is None).
    border : dict, optional
        Dictionary specifying border attributes like 'top_style', 'top_color', etc.
        (default is None).
    fill : dict, optional
        Dictionary specifying fill attributes like 'type' and 'color' (default is None).
    alignment : dict, optional
        Dictionary specifying alignment attributes such as 'horizontal',
        'vertical', and 'wrap_text' (default is None).
    number_format : str, optional
        The format string for cell numbers (default is "General").
    merge : bool, optional
        Whether the cell style supports merging (default is False).
    """

    def __init__(
        self,
        name: str,
        font: Optional[Dict[str, Union[str, bool, int]]] = None,
        border: Optional[Dict[str, Union[str, _Color]]] = None,
        fill: Optional[Dict[str, Union[str, _Color]]] = None,
        alignment: Optional[Dict[str, Union[str, bool, int]]] = None,
        number_format: Optional[str] = None,
        merge: bool = False,
    ):
        self.style = _NamedStyle(name=name or "-")
        self.name = name or "-"
        self.number_format = number_format or "General"
        self.merge = merge

        self.__set_font(font)
        self.__set_border(border)
        self.__set_fill(fill)
        self.__set_alignment(alignment)

    def __set_font(self, font: Optional[Dict[str, Union[str, bool, int]]]) -> None:
        """
        Configures the font settings for the style.

        Parameters
        ----------
        font : dict[str, Union[str, bool, int]] | None
            A dictionary specifying font properties such as:
            - 'name' (str): Font name (e.g., "Calibri").
            - 'bold' (bool): Whether the font is bold.
            - 'italic' (bool): Whether the font is italicized.
            - 'size' (int): Font size.
            - 'color' (str): Hexadecimal color code.
        """

        default_font = {
            "name": "Calibri",
            "bold": False,
            "italic": False,
            "size": 11,
            "color": "000000",
        }
        if font:
            default_font.update(font)
        self.style.font = _Font(
            name=default_font["name"],
            bold=default_font["bold"],
            italic=default_font["italic"],
            size=default_font["size"],
            color=_Color(default_font["color"]),
        )

    def __set_border(self, border: Optional[Dict[str, Union[str, _Color]]]) -> None:
        """
        Configures the border settings for the style.

        Parameters
        ----------
        border : dict[str, Union[str, _Color]] | None
            A dictionary specifying border properties for each side:
            - 'top_style', 'bottom_style', 'left_style', 'right_style' (str):
            Border styles (e.g., 'thin', 'medium').
            - 'top_color', 'bottom_color', 'left_color', 'right_color' (str):
            Hexadecimal color codes for each border.
        """

        default_border = {
            "top_style": None,
            "top_color": "000000",
            "bottom_style": None,
            "bottom_color": "000000",
            "left_style": None,
            "left_color": "000000",
            "right_style": None,
            "right_color": "000000",
        }
        if border:
            default_border.update(border)
        self.style.border = _Border(
            top=_Side(
                border_style=default_border["top_style"],
                color=_Color(default_border["top_color"]),
            ),
            bottom=_Side(
                border_style=default_border["bottom_style"],
                color=_Color(default_border["bottom_color"]),
            ),
            left=_Side(
                border_style=default_border["left_style"],
                color=_Color(default_border["left_color"]),
            ),
            right=_Side(
                border_style=default_border["right_style"],
                color=_Color(default_border["right_color"]),
            ),
        )

    def __set_fill(self, fill: Optional[Dict[str, Union[str, _Color]]]) -> None:
        """
        Configures the fill settings for the style.

        Parameters
        ----------
        fill : dict[str, Union[str, _Color]] | None
            A dictionary specifying fill properties:
            - 'type' (str): Fill pattern type (e.g., "solid", "none").
            - 'color' (str): Hexadecimal color code for the fill.
        """

        default_fill = {"type": "none", "color": "FFFFFF"}
        if fill:
            default_fill.update(fill)
        self.style.fill = _PatternFill(
            patternType=default_fill["type"], fgColor=_Color(default_fill["color"])
        )

    def __set_alignment(self, alignment: Optional[Dict[str, Union[str, bool, int]]]) -> None:
        """
        Configures the alignment settings for the style.

        Parameters
        ----------
        alignment : dict[str, Union[str, bool, int]] | None
            A dictionary specifying alignment properties:
            - 'horizontal' (str): Horizontal alignment (e.g., "left", "center").
            - 'vertical' (str): Vertical alignment (e.g., "top", "center").
            - 'wrap_text' (bool): Whether text should wrap.
            - 'text_rotation' (int): Text rotation angle (0-180).
            - 'shrink_to_fit' (bool): Whether text should shrink to fit the cell.
        """

        default_alignment = {
            "horizontal": "left",
            "vertical": "center",
            "wrap_text": False,
            "text_rotation": 0,
            "shrink_to_fit": False,
        }
        if alignment:
            default_alignment.update(alignment)
        self.style.alignment = _Alignment(
            horizontal=default_alignment["horizontal"],
            vertical=default_alignment["vertical"],
            wrap_text=default_alignment["wrap_text"],
            text_rotation=default_alignment["text_rotation"],
            shrink_to_fit=default_alignment["shrink_to_fit"],
        )


class Column:
    """
    Represents a column in an Excel sheet, including its title, content,
    styles, and formatting.

    Parameters
    ----------
    content : list of str
        The list of data entries in the column.
    title : str
        The column's title, which will appear in the first row.
    content_style : Style or list of Style, optional
        A single `Style` instance or a list of `Style` instances applied to
        the column content (default is None).
    title_style : Style, optional
        A `Style` instance applied to the column title (default is None).
    column_width : int, optional
        The width of the column (default is 11).
    best_fit : bool, optional
        If True, the column width adjusts automatically to fit content (default is False).
    """

    def __init__(
        self,
        content: List[str],
        title: str,
        content_style: Optional[Union[List[Style], Style]] = None,
        title_style: Optional[Style] = None,
        column_width: int = 11,
        best_fit: bool = False,
    ):
        self.content = content
        self.title = title
        self.content_style = self.__normalize_content_style(content, content_style)
        self.title_style = title_style or Style("")
        self.column_width = column_width
        self.best_fit = best_fit

    def __normalize_content_style(
        self,
        content: List[str],
        content_style: Optional[Union[List[Style], Style]],
    ) -> List[Style]:
        """
        Ensures the content style is a list of `Style` instances, matching the
        content length.

        Parameters
        ----------
        content : list of str
            The column's content, with one entry per row.
        content_style : Style or List[Style] | None
            A single `Style` instance, a list of `Style` instances, or None.

        Returns
        -------
        List[Style]
            A list of `Style` instances, with one entry per row.

        Raises
        ------
        CellmateException
            If a list of styles is provided but its length does not match the
            length of the content.
        """

        if content_style is None:
            return [Style("") for _ in content]
        elif isinstance(content_style, list):
            if len(content_style) != len(content):
                raise CellmateException(
                    "Content style length does not match content length"
                )
            return content_style
        else:
            return [content_style for _ in content]


class Sheet:
    """
    Creates an Excel sheet from formatted columns, applying styles to titles
    and content.

    Parameters
    ----------
    data : list of Column
        A list of `Column` instances defining the sheet's structure and content.
    sheet_name : str
        The name of the sheet.
    """

    def __init__(self, data: List[Column], sheet_name: str):
        self.sheet_name = sheet_name
        self.columns = data

    @property
    def named_styles(self) -> Set[_NamedStyle]:
        styles = []
        for col in self.columns:
            styles += [c.style for c in col.content_style]
            styles.append(col.title_style.style)

        return set(styles)

    def _populate_sheet(self, ws) -> None:
        """
        Populates the Excel sheet with data from the columns and applies the
        corresponding styles.

        Parameters
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet instance to populate.
        """

        for col_idx, col in enumerate(self.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col.column_width
            self.__cell_format(ws, 1, col_idx, col.title, col.title_style)
            for row_idx, value in enumerate(col.content, start=2):
                self.__cell_format(
                    ws, row_idx, col_idx, value, col.content_style[row_idx - 2]
                )

    def __cell_format(self, ws, row: int, col: int, value: str, style: Style) -> None:
        """
        Formats an individual cell with a value and style.

        Parameters
        ----------
        ws : openpyxl.worksheet.worksheet.Worksheet
            The worksheet instance containing the cell.
        row : int
            The row number of the cell.
        col : int
            The column number of the cell.
        value : str
            The value to assign to the cell.
        style : Style
            The `Style` instance to apply to the cell.
        """

        cell = ws.cell(row=row, column=col, value=value)
        cell.style = style.name
        cell.number_format = style.number_format

    def save(self, path: str) -> None:
        """Save sheet as a Workbook.

        Parameters
        ----------
        path : str
            File path where the sheet will be saved as a Workbook.

        """

        wb = _Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        for named_style in self.named_styles:
            wb.add_named_style(named_style)

        self._populate_sheet(ws)
        wb.save(path)


class Workbook(_Workbook):
    """
    Extends the openpyxl `Workbook` class to manage multiple sheets and styles.

    Parameters
    ----------
    sheets : list of Sheet
        A list of `Sheet` instances to include in the workbook.
    """

    def __init__(self, sheets: List[Sheet]) -> None:
        super().__init__()
        self.__add_named_styles(sheets)
        self.__insert_sheets(sheets)

    def __add_named_styles(self, sheets: List[Sheet]) -> None:
        """
        Adds all unique named styles from the sheets to the workbook.

        Parameters
        ----------
        sheets : list of Sheet
            A list of `Sheet` instances from which to extract styles.
        """

        styles = set()
        for sheet in sheets:
            styles.update(sheet.named_styles)

        for style in styles:
            self.add_named_style(style)

    def __insert_sheets(self, sheets: List[Sheet]) -> None:
        """
        Inserts all sheets into the workbook and applies their content and styles.

        Parameters
        ----------
        sheets : list of Sheet
            A list of `Sheet` instances to insert into the workbook.
        """

        for sheet in sheets:
            ws = self.create_sheet(sheet.sheet_name)
            sheet._populate_sheet(ws)

        # Remove first empty sheet
        ws = self.worksheets[0]
        self.remove(ws)
