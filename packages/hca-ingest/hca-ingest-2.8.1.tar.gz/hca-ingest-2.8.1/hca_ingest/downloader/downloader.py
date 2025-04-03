from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Protection, Font

from hca_ingest.importer.spreadsheet.ingest_workbook import SCHEMAS_WORKSHEET
from hca_ingest.importer.spreadsheet.ingest_worksheet import START_DATA_ROW

TITLE_ROW_NO = 1
TITLE_FONT = Font(bold=True, size=12)
TITLE_FILL = PatternFill(fill_type="solid", start_color='D0D0D0')
TITLE_ALIGNMENT = Alignment(vertical='center', wrap_text=True)
DESCRIPTION_ROW_NO = 2
DESCRIPTION_FONT = Font(italic=True, size=12, color='808080')
DESCRIPTION_ALIGNMENT = Alignment(vertical='top', wrap_text=True)
GUIDE_ROW_NO = 3
HEADER_ROW_NO = 4
HEADER_PROTECTION = Protection(locked=True)
BORDER_ROW_NO = 5
BORDER_VALUE = 'FILL OUT INFORMATION BELOW THIS ROW'

class XlsDownloader:
    @staticmethod
    def create_workbook(input_json: dict) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)

        for ws_title, ws_elements in input_json.items():
            if ws_title == 'Project':
                worksheet: Worksheet = workbook.create_sheet(title=ws_title, index=0)
            elif ws_title == SCHEMAS_WORKSHEET:
                continue
            else:
                worksheet: Worksheet = workbook.create_sheet(title=ws_title)
            XlsDownloader.add_worksheet_content(worksheet, ws_elements)

        XlsDownloader.generate_schemas_worksheet(input_json, workbook)
        return workbook

    @staticmethod
    def generate_schemas_worksheet(input_json: dict, workbook: Workbook):
        schemas = input_json.get(SCHEMAS_WORKSHEET)
        if not schemas:
            raise ValueError('The schema urls are missing')
        schemas_worksheet = workbook.create_sheet(SCHEMAS_WORKSHEET)
        schemas_worksheet.cell(row=1, column=1, value=SCHEMAS_WORKSHEET)
        for row_num, schema in enumerate(schemas, start=2):
            schemas_worksheet.cell(row=row_num, column=1, value=schema)

    @staticmethod
    def add_worksheet_content(worksheet: Worksheet, ws_elements: dict):
        headers = ws_elements.get('headers', {})
        XlsDownloader.__add_header_rows(worksheet, headers)
        all_values = ws_elements.get('values', [])

        for row_number, row_values in enumerate(all_values, start=START_DATA_ROW):
            XlsDownloader.__add_row_content(worksheet, headers, row_number, row_values)

    @staticmethod
    def __add_header_rows(worksheet: Worksheet, headers: dict):
        XlsDownloader.__add_border_row(worksheet)
        for col, header in enumerate(headers.keys(), start=1):
            XlsDownloader.__add_column_header(worksheet, col, header, headers.get(header, {}))

    @staticmethod
    def __add_column_header(worksheet: Worksheet, column_number: int, column_key: str, header_info: dict):
        title_cell: Cell = XlsDownloader.__add_title_header(
            worksheet.cell(row=TITLE_ROW_NO, column=column_number),
            header_info.get('user_friendly', ''),
            header_info.get('required', False)
        )
        XlsDownloader.__add_description_header(
            worksheet.cell(row=DESCRIPTION_ROW_NO, column=column_number),
            header_info.get('description', '')
        )
        XlsDownloader.__add_guide_header(
            worksheet.cell(row=GUIDE_ROW_NO, column=column_number),
            header_info.get('guidelines', ''),
            header_info.get('example', '')
        )
        XlsDownloader.__add_key_header(
            worksheet.cell(row=HEADER_ROW_NO, column=column_number),
            column_key
        )
        worksheet.column_dimensions[title_cell.column_letter].width = max(len(title_cell.value),len(column_key))

    @staticmethod
    def __add_title_header(cell: Cell, user_friendly: str = '', required: bool = False):
        value = user_friendly
        if required:
            value = f'{value} (Required)'
        cell.value = value
        cell.font = TITLE_FONT
        cell.fill = TITLE_FILL
        cell.alignment = TITLE_ALIGNMENT
        return cell

    @staticmethod
    def __add_description_header(cell: Cell, description: str):
        cell.value = description
        cell.font = DESCRIPTION_FONT
        cell.alignment = DESCRIPTION_ALIGNMENT

    @staticmethod
    def __add_guide_header(cell: Cell, guidelines: str, example: str):
        descriptions = []
        if guidelines:
            descriptions.append(guidelines)
        if example:
            descriptions.append(f'For example: {example}')
        description = ' '.join(descriptions)
        XlsDownloader.__add_description_header(cell, description)

    @staticmethod
    def __add_key_header(cell: Cell, column_key: str):
        cell.value = column_key
        cell.protection = HEADER_PROTECTION

    @staticmethod
    def __add_border_row(worksheet: Worksheet):
        # https://openpyxl.readthedocs.io/en/latest/styles.html#applying-styles
        # Applying a style to a row only applies it to cells created in that row in Excel
        row = worksheet.row_dimensions[BORDER_ROW_NO]
        row.font = TITLE_FONT
        row.fill = TITLE_FILL

        cell: Cell = worksheet.cell(row=BORDER_ROW_NO, column=1)
        cell.value = BORDER_VALUE
        cell.font = TITLE_FONT
        cell.fill = TITLE_FILL

    @staticmethod
    def __add_row_content(worksheet: Worksheet, headers: dict, row_number: int, values: dict):
        for index, header in enumerate(headers.keys(), start=1):
            if header in values:
                worksheet.cell(row=row_number, column=index, value=values[header])
