import openpyxl as op

from ka_uts_dic.dic import Dic
from ka_uts_obj.path import Path
from ka_uts_df.pddf import PdDf

import pandas as pd

from typing import Any, Dict, List

T_Arr = List[Any]
T_Dic = Dict[Any, Any]
T_StrArr = str | T_Arr
TN_Arr = None | T_Arr
TN_Dic = None | T_Dic
TN_Int = None | int


class DoPdDf:

    @staticmethod
    def write_xlsx(d_df: T_Dic, path: str, **kwargs) -> None:
        Path.mkdir_from_path(path)
        writer = pd.ExcelWriter(path, engine='openpyxl')
        _a_key: T_Arr = Dic.show_sorted_keys(d_df)
        for _key in _a_key:
            _df = d_df[_key]
            _df.to_excel(writer, sheet_name=_key)
        writer.close()

    @staticmethod
    def update_xlsx_cell_with_d_body(
            ws_new, df: pd.DataFrame, d_df2xlsx: T_Dic) -> None:
        _d_update: T_Dic = d_df2xlsx.get('update', {})
        _pv_indexes: T_StrArr = _d_update.get('pv_indexes', [])
        _pv_a_nm_col: T_Dic = _d_update.get('pv_a_nm_col', {})
        _d_body: T_Dic = _d_update.get('d_body', {})
        _a_nm_row: T_Arr = _d_body.get('a_row', [])
        _a_nm_col: T_Arr = _d_body.get('a_col', [])
        _xlsx_row_offset: int = Dic.locate(_d_body, ['offset', 'row'])
        _xlsx_col_offset: int = Dic.locate(_d_body, ['offset', 'col'])

        df_new = df.set_index(_pv_indexes)
        a_df_nm_row = df_new.index
        print(f"update_xlsx_cell_with_d_body df = {df}")
        print(f"update_xlsx_cell_with_d_body a_df_nm_row = {a_df_nm_row}")

        for _nm_row in a_df_nm_row:
            _ix_row = _a_nm_row.index(_nm_row)+_xlsx_row_offset
            for _nm_col in _pv_a_nm_col:
                print(f"_nm_row={_nm_row}")
                print(f"_nm_col={_nm_col}")
                # print(f"_ix_row={_ix_row}")
                # print(f"_ix_col={_ix_col}")
                _df_cell_value = df_new.loc[_nm_row, _nm_col]
                # print(f"_df_cell_value = {_df_cell_value}")
                _ix_col = _a_nm_col.index(_nm_col)+_xlsx_col_offset
                ws_new.cell(_ix_row, _ix_col).value = _df_cell_value

    @staticmethod
    def update_xlsx_cell_with_d_head(ws_new, d_head) -> None:
        for _key in d_head.keys():
            _ix_row = d_head[_key]['row']
            _ix_col = d_head[_key]['col']
            _value = d_head[_key]['value']
            ws_new.cell(_ix_row, _ix_col).value = _value

    @classmethod
    def update_xlsx(cls, d_df: T_Dic, **kwargs):
        _d_update: T_Dic = kwargs.get('d_update', {})
        _d_head: T_Dic = _d_update.get('d_head', {})

        _a_key: T_Arr = Dic.show_sorted_keys(d_df)

        _dl_in_path_tmpl_rpt = kwargs.get("dl_in_path_tmpl_rpt", '')
        _wb_tmpl = op.load_workbook(filename=_dl_in_path_tmpl_rpt)

        for _key in _a_key:
            _df = d_df[_key]
            _ws_tmpl = _wb_tmpl['TMPL']
            _ws_new = _wb_tmpl.copy_worksheet(_ws_tmpl)
            _ws_new.title = _key
            _d_head['title']['value'] = _key
            cls.update_xlsx_cell_with_d_body(_ws_new, _df, _d_update)
            cls.update_xlsx_cell_with_d_head(_ws_new, _d_head)
        return _wb_tmpl

    @classmethod
    def ioc_xlsx(
            cls, d_df: T_Dic, path: str, **kwargs) -> None:
        print("ioc_xlsx START")
        _wb_tmpl = cls.update_xlsx(d_df, **kwargs)
        Path.mkdir_from_path(path)
        print(f"ioc_xlsx END path = {path}")
        print(f"ioc_xlsx END _wb_tmpl = {_wb_tmpl}")
        _wb_tmpl.save(path)

    @staticmethod
    def set_index_drop_key_filter(
            d_df: T_Dic, d_filter: T_Dic, relation: str, index: str) -> T_Dic:
        _a_key: T_Arr = Dic.show_sorted_keys(d_df)
        for _key in _a_key:
            df_new = d_df[_key]
            df_new = PdDf.set_index_drop_column_filter(
                    df_new, d_filter, relation, index)
            d_df[_key] = df_new
        return d_df
