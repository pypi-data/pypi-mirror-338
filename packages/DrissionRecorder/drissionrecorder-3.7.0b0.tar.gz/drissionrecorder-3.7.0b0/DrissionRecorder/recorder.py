# -*- coding:utf-8 -*-
from pathlib import Path
from time import sleep

from openpyxl.utils import column_index_from_string

from .base import BaseRecorder
from .setter import RecorderSetter, set_csv_header
from .style.cell_style import CellStyleCopier, CellStyle
from .tools import (ok_list, data_to_list_or_dict, process_content, data_to_list_or_dict_simplify, get_wb,
                    get_and_set_csv_header, get_ws, create_csv, get_header, fix_openpyxl_bug, handle_new_sheet)


class Recorder(BaseRecorder):
    _SUPPORTS = ('csv', 'xlsx', 'txt', 'jsonl', 'json')

    def __init__(self, path=None, cache_size=None):
        """用于缓存并记录数据，可在达到一定数量时自动记录，以降低文件读写次数，减少开销
        :param path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件，0为不自动写入
        """
        super().__init__(path=path, cache_size=cache_size)
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符
        self._follow_styles = False
        self._col_height = None
        self._styles = None
        self._fit_header = False
        self._header_row = 1
        self._auto_new_col = False

    @property
    def set(self):
        """返回用于设置属性的对象"""
        if self._setter is None:
            self._setter = RecorderSetter(self)
        return self._setter

    @property
    def delimiter(self):
        """返回csv文件分隔符"""
        return self._delimiter

    @property
    def quote_char(self):
        """返回csv文件引用符"""
        return self._quote_char

    @property
    def header(self):
        """返回表头，只支持csv和xlsx格式"""
        if self.type not in ('csv', 'xlsx'):
            raise TypeError('header属性只支持csv和xlsx类型文件。')
        return get_header(self)

    def key2col(self, key, table=None):
        """接收关键字，返回该关键字在表头中的序号
        :param key: 关键字
        :param table: xlsx格式时可指定在某个sheet中获取，为None时使用对象设置的值
        :return: 序号，从1开始，没有该关键字时报错
        """
        header = get_header(self, table=table)
        if header and key in header:
            return header.index(key) + 1
        raise RuntimeError(f'没有找到指定表格或表头中没有该值：{key}')

    def add_data(self, data, table=None):
        """添加数据，可一次添加多条数据
        :param data: 插入的数据，任意格式
        :param table: 要写入的数据表，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        while self._pause_add:  # 等待其它线程写入结束
            sleep(.1)

        if not isinstance(data, (list, tuple, dict)):
            data = (data,)

        if not data:
            data = ([],)
            self._data_count += 1

        # 一维数组
        elif isinstance(data, dict) or (isinstance(data, (list, tuple))
                                        and not isinstance(data[0], (list, tuple, dict))):
            data = [data_to_list_or_dict(self, data)]
            self._data_count += 1

        else:  # 二维数组
            if self.after or self.before:
                data = [data_to_list_or_dict(self, d) for d in data]
            else:
                data = [data_to_list_or_dict_simplify(d) for d in data]
            self._data_count += len(data)

        if self._type != 'xlsx':
            self._data.extend(data)

        else:
            if table is None:
                table = self._table
            elif isinstance(table, bool):
                table = None

            self._data.setdefault(table, []).extend(data)

        if 0 < self.cache_size <= self._data_count:
            self.record()

    def _record(self):
        """记录数据"""
        if self.type == 'csv':
            self._to_csv()
        elif self.type == 'xlsx':
            self._to_xlsx()
        elif self.type == 'txt':
            self._to_txt()
        elif self.type == 'jsonl':
            self._to_jsonl()
        elif self.type == 'json':
            self._to_json()

    def _to_xlsx(self):
        """记录数据到xlsx文件"""
        wb, new_file = get_wb(self)
        tables = [i.title for i in wb.worksheets]

        for table, data in self._data.items():
            _row_styles = None
            _col_height = None
            ws, new_sheet = get_ws(wb, table, tables, new_file)
            first_wrote = False

            # ==============处理表头和样式==============
            if new_sheet:
                first_wrote = handle_new_sheet(self, ws, data, first_wrote)

            elif self._header_row > 0 and not self._header.get(ws.title, None):
                self._header[ws.title] = [c.value for c in ws[self._header_row]]

            begin_row = None  # 开始写入数据的行
            if self._follow_styles:
                begin_row = ws.max_row
                _row_styles = [CellStyleCopier(i) for i in ws[begin_row]]
                _col_height = ws.row_dimensions[begin_row].height
                begin_row += 1
            elif self._styles or self._col_height:
                begin_row = ws.max_row + 1

            if new_file:
                wb, ws = fix_openpyxl_bug(self, wb, ws, ws.title)
            new_file = False

            # ==============开始写入数据==============
            if first_wrote:
                data = data[1:]
            if self._fit_header and self._header.get(ws.title, None):
                rewrite_header = False
                for i in data:
                    i, rewrite_header = _fit_header_handle(i, self, ws, rewrite_header)
                    ws.append(ok_list(i, True))

                if rewrite_header:
                    for c in ws[1]:
                        c.value = None
                    for k, i in enumerate(self._header[ws.title], 1):
                        ws.cell(1, k).value = i

            elif self._fit_header and self._header_row == 0:
                for i in data:
                    if isinstance(i, dict):
                        d = {}
                        for k, v in i.items():
                            try:
                                d[column_index_from_string(k)] = v
                            except ValueError:
                                pass
                        d = list(d.items())
                        d.sort(key=lambda x: x[0])
                        t = []
                        wei = 0
                        for k, v in d:
                            for j in range(k - wei - 1):
                                t.append(None)
                            t.append(v)
                            wei = k
                        i = t
                    ws.append(ok_list(i, True))

            else:
                for i in data:
                    ws.append(ok_list(i, True))

            if self._follow_styles:
                for r in range(begin_row, ws.max_row + 1):
                    _set_style(_col_height, _row_styles, ws, r)

            elif self._styles or self._col_height:
                for r in range(begin_row, ws.max_row + 1):
                    _set_line_style(self._col_height, self._styles, ws, r, ws.max_column)

        wb.save(self.path)
        wb.close()

    def _to_csv(self):
        """记录数据到csv文件"""
        if self._header is not None and not self._file_exists:
            create_csv(self)
        elif self._header is None and self._header_row > 0:
            get_and_set_csv_header(self)

        rewrite_header = False
        with open(self.path, 'a+', newline='', encoding=self.encoding) as f:
            from csv import writer
            csv_write = writer(f, delimiter=self.delimiter, quotechar=self.quote_char)
            if self._fit_header and self._header:
                method = _to_csv1
            elif self._fit_header and self._header_row == 0:
                method = _to_csv2
            else:
                method = _to_csv3

            if method(self, csv_write, rewrite_header):
                set_csv_header(self, self._header, True, self._header_row)

    def _to_txt(self):
        """记录数据到txt文件"""
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = [' '.join(ok_list(i, as_str=True)) for i in self._data]
            f.write('\n'.join(all_data) + '\n')

    def _to_jsonl(self):
        """记录数据到jsonl文件"""
        from json import dumps
        with open(self.path, 'a+', encoding=self.encoding) as f:
            all_data = [i if isinstance(i, str) else dumps(i) for i in self._data]
            f.write('\n'.join(all_data) + '\n')

    def _to_json(self):
        """记录数据到json文件"""
        from json import load, dump
        if self._file_exists or Path(self.path).exists():
            with open(self.path, 'r', encoding=self.encoding) as f:
                json_data = load(f)
        else:
            json_data = []

        for i in self._data:
            if isinstance(i, dict):
                for d in i:
                    i[d] = process_content(i[d])
                json_data.append(i)
            else:
                json_data.append([process_content(d) for d in i])

        self._file_exists = True
        with open(self.path, 'w', encoding=self.encoding) as f:
            dump(json_data, f)


def _set_style(height, styles, ws, row):
    if height is not None:
        ws.row_dimensions[row].height = height

    if styles:
        if isinstance(styles, CellStyle):
            for c in ws[row]:
                styles.to_cell(c)
        else:
            for k, s in enumerate(styles, start=1):
                if s:
                    s.to_cell(ws.cell(row=row, column=k))


def _set_line_style(height, style, ws, row, max_col):
    if height is not None:
        ws.row_dimensions[row].height = height
    if style:
        for i in range(1, max_col + 1):
            style.to_cell(ws.cell(row=row, column=i))


def _fit_header_handle(data, recorder, ws, rewrite_header):
    """处理需要匹配表头时数据"""
    if isinstance(data, dict):
        if recorder._auto_new_col and set(recorder._header[ws.title]) != set(data.keys()):
            recorder._header[ws.title] += [t for t in data.keys() if t not in recorder._header[ws.title]]
            rewrite_header = True
        data = [data.get(h, None) for h in recorder._header[ws.title]]
    return data, rewrite_header


def _style_handle(recorder, data):
    """处理需要匹配样式时数据"""
    return [recorder._style] * len(data) if isinstance(recorder._style, CellStyle) else recorder._style


def _to_csv1(recorder, writer, rewrite_header):
    for i in recorder._data:
        if isinstance(i, dict):
            if recorder._auto_new_col and set(recorder._header) != set(i.keys()):
                recorder._header += [t for t in i.keys() if t not in recorder._header]
                rewrite_header = True
            i = [i.get(h, '') for h in recorder._header]
        writer.writerow(ok_list(i))
    return rewrite_header


def _to_csv2(recorder, writer, rewrite_header):
    for i in recorder._data:
        if isinstance(i, dict):
            i = [(k, v) for k, v in i.items() if isinstance(k, int) and k > 0]
            i.sort(key=lambda x: x[1])
            t = []
            wei = 0
            for k, v in i:
                for j in range(k - wei - 1):
                    t.append('')
                t.append(str(v))
                wei = k
            i = t
        writer.writerow(ok_list(i))
    return rewrite_header


def _to_csv3(recorder, writer, rewrite_header):
    for i in recorder._data:
        writer.writerow(ok_list(i))
    return rewrite_header
