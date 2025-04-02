# -*- coding:utf-8 -*-
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook

from .tools import process_content, ok_list, make_valid_name, format_signs


class OriginalSetter(object):
    def __init__(self, recorder):
        self._recorder = recorder

    def cache_size(self, size):
        """设置缓存大小
        :param size: 缓存大小
        :return: 设置对象自己
        """
        if not isinstance(size, int) or size < 0:
            raise TypeError('cache_size值只能是int，且必须>=0')
        self._recorder._cache = size
        return self

    def path(self, path):
        """设置文件路径
        :param path: 文件路径
        :return: 设置对象自己
        """
        if self._recorder._path:
            self._recorder.record()

        p = Path(path)
        self._recorder._path = str(p.parent / make_valid_name(p.name))
        self._recorder._data = []
        self._recorder._file_exists = False
        return self

    def show_msg(self, on_off):
        """设置是否显示运行信息
        :param on_off: bool表示开关
        :return: 设置对象自己
        """
        self._recorder.show_msg = on_off
        return self

    def auto_backup(self, interval=None, path=None, new_name=None):
        """设置自动备份相关参数
        :param interval: 自动保存多少次时触发备份，为0表示不自动备份，为None时不修改已设置值（初始为0）
        :param path: 备份文件存放路径，为None时不修改已设置值（初始为 'backup'）
        :param new_name: 是否用新名称备份，为None时不修改已设置值（初始为True）
        :return: 设置对象自己
        """
        if path is not None:
            self._recorder._backup_path = path
        if isinstance(new_name, bool):
            self._recorder._backup_new_name = new_name
        if interval is not None:
            self._recorder._backup_interval = interval
        return self


class BaseSetter(OriginalSetter):
    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: 设置对象自己
        """
        self._recorder._table = name
        return self

    def before(self, before):
        """设置在数据前面补充的列
        :param before: 列表、元组或字符串，为字符串时则补充一列
        :return: 设置对象自己
        """
        if before is None:
            self._recorder._before = None
        elif isinstance(before, (list, dict)):
            self._recorder._before = before
        elif isinstance(before, tuple):
            self._recorder._before = list(before)
        else:
            self._recorder._before = [before]
        return self

    def after(self, after):
        """设置在数据后面补充的列
        :param after: 列表、元组或字符串，为字符串时则补充一列
        :return: 设置对象自己
        """
        if after is None:
            self._recorder._after = None
        elif isinstance(after, (list, dict)):
            self._recorder._after = after
        elif isinstance(after, tuple):
            self._recorder._after = list(after)
        else:
            self._recorder._after = [after]
        return self

    def encoding(self, encoding):
        """设置编码
        :param encoding: 编码格式
        :return: 设置对象自己
        """
        self._recorder._encoding = encoding
        return self


class SheetLikeSetter(BaseSetter):
    def header(self, header, table=None, to_file=True):
        """设置表头。只有 csv 和 xlsx 格式支持设置表头
        :param header: 表头，列表或元组
        :param table: 表名，只xlsx格式文件有效
        :param to_file: 是否写入到文件
        :return: 设置对象自己
        """
        self._recorder.record()
        with self._recorder._lock:
            if not self._recorder.path:
                raise FileNotFoundError('未指定文件。')
            if not isinstance(header, (list, tuple)):
                raise TypeError('header参数只能是list或tuple格式。')

            if self._recorder.type == 'xlsx':
                table = table or self._recorder.table
                set_xlsx_header(self._recorder, header, table, to_file, self._recorder._header_row)

            elif self._recorder.type == 'csv':
                set_csv_header(self._recorder, header, to_file, self._recorder._header_row)

            else:
                raise TypeError('只能对xlsx和csv文件设置表头。')
        return self

    def delimiter(self, delimiter):
        """设置csv文件分隔符
        :param delimiter: 分隔符
        :return: 设置对象自己
        """
        self._recorder._delimiter = delimiter
        return self

    def quote_char(self, quote_char):
        """设置csv文件引用符
        :param quote_char: 引用符
        :return: 设置对象自己
        """
        self._recorder._quote_char = quote_char
        return self

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: 设置对象自己
        """
        super().path(path)

        if not file_type:
            suffix = Path(path).suffix.lower()
            if suffix:
                file_type = suffix[1:]
            elif not self._recorder.type:
                file_type = 'csv'

        if file_type:
            self.file_type(file_type)

        if self._recorder._type == 'xlsx':
            self._recorder._data = {}
            self._recorder._header = {}
            self._recorder._style_data = {}
        else:
            self._recorder._data = []
            self._recorder._header = None

        return self

    def file_type(self, file_type):
        """指定文件类型，无视文件后缀名"""
        if file_type not in self._recorder._SUPPORTS:
            file_type = 'csv'
        self._recorder._type = file_type
        return self

    def table(self, name):
        """设置默认表名
        :param name: 表名，为None表示使用活动表格
        :return: 设置对象自己
        """
        if isinstance(name, bool):
            name = None
        self._recorder._table = name
        return self

    def header_row(self, num):
        """设置标题行号
        :param num: 行号
        :return: 设置对象自己
        """
        if num < 0:
            raise ValueError('num不能小于0。')
        self._recorder._header_row = num
        if num == 0:
            if self._recorder.type == 'csv':
                self._recorder._header = None
            else:
                self._recorder._header.pop(self._recorder.table, None)
        return self

    def fit_header(self, on_off=True):
        """设置是否自动匹配表头
        :param on_off: bool表示开关
        :return: 设置对象自己
        """
        if self._recorder.type not in ('csv', 'xlsx'):
            raise TypeError('只有csv或xlsx格式可设置表头。')
        self._recorder.record()
        self._recorder._fit_header = on_off
        return self

    def follow_styles(self, on_off=True):
        """设置是否跟随上一行的style，只有xlsx格式有效
        :param on_off: True或False
        :return: 设置对象自己
        """
        self._recorder._follow_styles = on_off
        return self

    # ---------即将废弃------------
    def head_row(self, num):
        return self.header_row(num)

    def head(self, head, table=None, to_file=True):
        return self.header(head, table, to_file)

    def fit_head(self, on_off=True):
        return self.fit_header(on_off)


class FillerSetter(SheetLikeSetter):
    def signs(self, values, deny=False):
        """设置sign值
        :param values: 筛选条件，可用list, tuple, set设置多个
        :param deny: 是否匹配非指定条件的数据
        :return: 设置对象自己
        """
        self._recorder._sign = format_signs(values)
        self._recorder._deny_sign = deny
        return self

    def key_cols(self, cols):
        """设置作为关键字的列，可以是多列
        :param cols: 列号或列名，或它们组成的list或tuple
        :return: 设置对象自己
        """
        self._recorder._key_cols = get_key_cols(cols)
        return self

    def sign_col(self, col):
        """设置用于判断是否已填数据的列
        :param col: 列号或列名
        :return: 设置对象自己
        """
        self._recorder._sign_col = get_sign_col(col)
        return self

    def data_col(self, col):
        """设置用于填充数据的列
        :param col: 列号或列名
        :return: 设置对象自己
        """
        if isinstance(col, int) and col > 0:
            self._recorder._data_col = col
        elif isinstance(col, str):
            self._recorder._data_col = column_index_from_string(col)
        else:
            raise TypeError('col值只能是int或str，且必须大于0。')
        return self

    def begin_row(self, row):
        """设置数据开始的行
        :param row: 行号
        :return: 设置对象自己
        """
        if not isinstance(row, int) or row < 1:
            raise TypeError('row值只能是int，且必须大于0')
        self._recorder._begin_row = row
        return self

    def link_style(self, style):
        """设置单元格的链接样式
        :param style: CellStyle对象
        :return: 设置对象自己
        """
        self._recorder._link_style = style
        return self

    # ------------ 即将废弃 ------------
    def deny_sign(self, on_off=True):
        """设置是否反向匹配sign
        :param on_off: bool表示开或关
        :return: 设置对象自己
        """
        self._recorder._deny_sign = on_off
        return self

    def sign(self, value, deny=False):
        """设置sign值
        :param value: 筛选条件
        :param deny: 是否匹配非指定条件的数据
        :return: 设置对象自己
        """
        return self.signs(value, deny)


class RecorderSetter(SheetLikeSetter):

    def col_height(self, height):
        """设置行高，只有xlsx格式有效
        :param height: 行高，传入None清空设置
        :return: 设置对象自己
        """
        self._recorder._col_height = height
        return self

    def styles(self, styles):
        """设置新行样式，只有xlsx格式有效，可传入多个，传入None则取消
        :param styles: 传入CellStyle对象设置整个新行，传入CellStyle对象组成的列表设置多个，传入None清空设置
        :return: 设置对象自己
        """
        self._recorder.record()
        self._recorder._follow_styles = False
        self._recorder._styles = styles
        return self

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: 设置对象自己
        """
        super().path(path=path, file_type=file_type)
        self._recorder._row_styles = None
        return self

    def fit_header(self, on_off=True, add_new=False):
        """设置是否自动匹配表头
        :param on_off: bool表示开关
        :param add_new: 数据中有表头不存在的列时是否自动添加到表头，on_off为True时才有效
        :return: 设置对象自己
        """
        super().fit_header(on_off)
        self._recorder._auto_new_col = add_new
        return self

    def file_type(self, file_type):
        """指定文件类型，无视文件后缀名"""
        if file_type not in self._recorder._SUPPORTS:
            file_type = 'txt'
        self._recorder._type = file_type
        return self

    # ---------即将废弃------------
    def fit_head(self, on_off=True, add_new=False):
        return self.fit_header(on_off, add_new)


class DBSetter(BaseSetter):
    def path(self, path, table=None):
        """重写父类方法
        :param path: 文件路径
        :param table: 数据表名称
        :return: 设置对象自己
        """
        with self._recorder._lock:
            super().path(path)
            if self._recorder._conn is not None:
                self._recorder._close_connection()
            self._recorder._connect()

            if table:
                self.table(table)
            else:
                r = self._recorder.run_sql("select name from sqlite_master where type='table'")
                self._recorder._table = r[0] if r else None

            self._recorder._data = {}
            self._recorder._close_connection()
        return self

    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: 设置对象自己
        """
        if '`' in name:
            raise ValueError('table名称不能包含字符"`"。')
        self._recorder._table = name
        return self


def set_csv_header(recorder, header, to_file, row=1):
    """设置csv文件的表头
    :param recorder: Recorder或Filler对象
    :param header: 表头列表或元组
    :param to_file: 是否写入文件
    :param row: 行号
    :return: None
    """
    recorder._header = header
    if not to_file:
        return

    from csv import writer
    if recorder._file_exists or Path(recorder.path).exists():
        with open(recorder.path, 'r', newline='', encoding=recorder._encoding) as f:
            lines = f.readlines()
            content1 = lines[:row - 1]
            content2 = lines[row:]

        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            f.write("".join(content1))
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            con_len = len(content1)
            if con_len < row - 1:
                for _ in range(row - con_len - 1):
                    csv_write.writerow([])
            csv_write.writerow(ok_list(header))

        with open(recorder.path, 'a+', newline='', encoding=recorder._encoding) as f:
            f.write("".join(content2))

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            for _ in range(row - 1):
                csv_write.writerow([])
            csv_write.writerow(ok_list(header))


def set_xlsx_header(recorder, header, table, to_file, row=1):
    """设置xlsx文件的表头
    :param recorder: Recorder或Filler对象
    :param header: 表头列表或元组
    :param table: 工作表名称
    :param to_file: 是否写入文件
    :param row: 行号
    :return: None
    """
    if not to_file:
        if table:
            recorder._header[table] = header
        elif recorder._file_exists or Path(recorder.path).exists():
            wb = load_workbook(recorder.path)
            ws = wb.active
            recorder._header[ws.title] = header
            wb.close()
        else:
            recorder._header['Sheet'] = header
        return

    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        if table:
            ws = wb[table] if table in [i.title for i in wb.worksheets] else wb.create_sheet(title=table)
        else:
            ws = wb.active

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        if table:
            ws.title = table

    if len(ws[row]) > len(header):
        header = list(header)
        header.extend([None] * (len(ws[row]) - len(header)))

    for key, i in enumerate(header, 1):
        ws.cell(row, key).value = process_content(i, True)

    recorder._header[ws.title] = header
    wb.save(recorder.path)
    wb.close()


def get_key_cols(cols):
    """获取作为关键字的列，可以是多列
    :param cols: 列号或列名，或它们组成的list或tuple
    :return: 列序号列表
    """
    if cols is True:
        return True
    elif isinstance(cols, int) and cols > 0:
        return [cols]
    elif isinstance(cols, str):
        return [int(cols)] if cols.isdigit() else [column_index_from_string(cols)]
    elif isinstance(cols, (list, tuple)):
        return [i if isinstance(i, int) and i > 0 else
                int(i) if i.isdigit() else column_index_from_string(i) for i in cols]
    else:
        raise TypeError('col值只能是int或str，且必须大于0。')


def get_sign_col(col):
    """设置用于判断是否已填数据的列
    :param col: 列号或列名
    :return: 列序号
    """
    if col is True or (isinstance(col, int) and col > 0):
        return col
    elif isinstance(col, str):
        return int(col) if col.isdigit() else column_index_from_string(col)
    else:
        raise TypeError('col值只能是True、int或str，且必须大于0。')
