# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from re import search, sub, match

from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook import Workbook

from .style import CellStyle


def align_csv(path, encoding='utf-8', delimiter=',', quotechar='"'):
    """补全csv文件，使其每行列数一样多，用于pandas读取时避免出错
    :param path: 要处理的文件路径
    :param encoding: 文件编码
    :param delimiter: 分隔符
    :param quotechar: 引用符
    :return: None
    """
    with open(path, 'r', encoding=encoding) as f:
        reader = csv_reader(f, delimiter=delimiter, quotechar=quotechar)
        lines = list(reader)
        lines_data = {}
        max_len = 0

        # 把每行列数用字典记录，并找到最长的一行
        for k, i in enumerate(lines):
            line_len = len(i)
            if line_len > max_len:
                max_len = line_len
            lines_data[k] = line_len

        # 把所有行用空值补全到和最长一行一样
        for i in lines_data:
            lines[i].extend([None] * (max_len - lines_data[i]))

        writer = csv_writer(open(path, 'w', encoding=encoding, newline=''), delimiter=delimiter, quotechar=quotechar)
        writer.writerows(lines)


def get_usable_path(path, is_file=True, parents=True):
    """检查文件或文件夹是否有重名，并返回可以使用的路径
    :param path: 文件或文件夹路径
    :param is_file: 目标是文件还是文件夹
    :param parents: 是否创建目标路径
    :return: 可用的路径，Path对象
    """
    path = Path(path)
    parent = path.parent
    if parents:
        parent.mkdir(parents=True, exist_ok=True)
    path = parent / make_valid_name(path.name)
    name = path.stem if path.is_file() else path.name
    ext = path.suffix if path.is_file() else ''

    first_time = True

    while path.exists() and path.is_file() == is_file:
        r = search(r'(.*)_(\d+)$', name)

        if not r or (r and first_time):
            src_name, num = name, '1'
        else:
            src_name, num = r.group(1), int(r.group(2)) + 1

        name = f'{src_name}_{num}'
        path = parent / f'{name}{ext}'
        first_time = None

    return path


def make_valid_name(full_name):
    """获取有效的文件名
    :param full_name: 文件名
    :return: 可用的文件名
    """
    # ----------------去除前后空格----------------
    full_name = full_name.strip()

    # ----------------去除不允许存在的字符----------------
    if search(r'[<>/\\|:*?\n"]', full_name):
        full_name = sub(r'<', '＜', full_name)
        full_name = sub(r'>', '＞', full_name)
        full_name = sub(r'/', '／', full_name)
        full_name = sub(r'\\', '＼', full_name)
        full_name = sub(r'\|', '｜', full_name)
        full_name = sub(r':', '：', full_name)
        full_name = sub(r'\*', '＊', full_name)
        full_name = sub(r'\?', '？', full_name)
        full_name = sub(r'\n', '', full_name)
        full_name = sub(r'"(.*?)"', r'“\1”', full_name)
        full_name = sub(r'"', '“', full_name)

    # ----------------使总长度不大于255个字符（一个汉字是2个字符）----------------
    r = search(r'(.*)(\.[^.]+$)', full_name)  # 拆分文件名和后缀名
    if r:
        name, ext = r.group(1), r.group(2)
        ext_long = len(ext)
    else:
        name, ext = full_name, ''
        ext_long = 0

    while get_long(name) > 255 - ext_long:
        name = name[:-1]

    return f'{name}{ext}'.rstrip('.')


def get_long(txt):
    """返回字符串中字符个数（一个汉字是2个字符）
    :param txt: 字符串
    :return: 字符个数
    """
    txt_len = len(txt)
    return int((len(txt.encode('utf-8')) - txt_len) / 2 + txt_len)


def parse_coord(coord=None, data_col=None):
    """添加数据，每次添加一行数据，可指定坐标、列号或行号
    coord只输入数字（行号）时，列号为self.data_col值，如 3；
    输入列号，或没有行号的坐标时，表示新增一行，列号为此时指定的，如'c'、',3'、(None, 3)、'None,3'；
    输入 'newline' 时，表示新增一行，列号为self.data_col值；
    输入行列坐标时，填写到该坐标，如'a3'、'3,1'、(3,1)、[3,1]；
    输入的行号可以是负数（列号不可以），代表从下往上数，-1是倒数第一行，如'a-3'、(-3, 3)
    :param coord: 坐标、列号、行号
    :param data_col: 列号，用于只传入行号的情况
    :return: 坐标tuple：(行, 列)，或(None, 列)
    """
    return_coord = None
    if coord == 'newline':  # 新增一行，列为data_col
        return_coord = None, data_col

    elif isinstance(coord, (int, float)) and coord != 0:
        return_coord = int(coord), data_col

    elif isinstance(coord, str):
        coord = coord.replace(' ', '')

        if coord.isalpha():  # 只输入列号，要新建一行
            return_coord = None, column_index_from_string(coord)

        elif ',' in coord:  # '3,1'形式
            x, y = coord.split(',')
            if x.lower() in ('', 'new', 'none', 'newline'):
                x = None
            elif x.isdigit():
                x = int(x)
            else:
                raise ValueError('行格式不正确。')

            if y.isdigit():
                y = int(y)
            elif y.isalpha():
                y = column_index_from_string(y)
            else:
                raise TypeError('列格式不正确。')

            return_coord = x, y

        else:  # 'A3'或'3A'形式
            m = match(r'^[$]?([A-Za-z]{1,3})[$]?(-?\d+)$', coord)
            if m:
                y, x = m.groups()
                return_coord = int(x), column_index_from_string(y)

            else:
                m = match(r'^[$]?(-?\d+)[$]?([A-Za-z]{1,3})$', coord)
                if not m:
                    raise ValueError(f'{coord} 坐标格式不正确。')
                x, y = m.groups()
                return_coord = int(x), column_index_from_string(y)

    elif isinstance(coord, (tuple, list)):
        if len(coord) != 2:
            raise ValueError('coord为list或tuple时长度必须为2。')

        x = None
        if coord[0] not in (None, 'new', 'newline'):
            x = int(coord[0])

        if isinstance(coord[1], int):
            y = coord[1]
        elif isinstance(coord[1], str):
            y = column_index_from_string(coord[1])
        else:
            raise TypeError('列格式不正确。')

        return_coord = x, y

    if not return_coord or return_coord[0] == 0 or return_coord[1] == 0:
        raise ValueError(f'{return_coord} 坐标格式不正确。')
    return return_coord


def process_content(content, excel=False):
    """处理单个单元格要写入的数据
    :param content: 未处理的数据内容
    :param excel: 是否为excel文件
    :return: 处理后的数据
    """
    if isinstance(content, (int, str, float, type(None))):
        data = content
    elif isinstance(content, (Cell, ReadOnlyCell)):
        data = content.value
    else:
        data = str(content)

    if excel and isinstance(data, str):
        data = sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', data)

    return data


def ok_list(data_list, excel=False, as_str=False):
    """处理列表中数据使其符合保存规范
    :param data_list: 数据列表
    :param excel: 是否保存在excel
    :param as_str: 内容是否转为字符串
    :return: 处理后的列表
    """
    if isinstance(data_list, dict):
        data_list = data_list.values()
    if as_str:
        data_list = [str(i) for i in data_list]
    return [process_content(i, excel) for i in data_list]


def get_usable_coord_int(coord, max_row, max_col):
    """返回真正写入文件的坐标
    :param coord: 已初步格式化的坐标，如(1, 2)、(None, 3)、(-3, -2)
    :param max_row: 文件最大行
    :param max_col: 文件最大列
    :return: 真正写入文件的坐标，tuple格式
    """
    row, col = coord
    if col < 0:
        col = max_col + col + 1
        if col < 1:
            raise ValueError(f'列号不能小于1。当前：{col}')

    if row is None:
        row = max_row + 1
    elif row < 0:
        row = max_row + row + 1
        if row < 1:
            raise ValueError(f'行号不能小于1。当前：{row}')

    return row, col


def get_usable_coord(coord, max_row, ws):
    """返回真正写入文件的坐标
    :param coord: 已初步格式化的坐标，如(1, 2)、(None, 3)、(-3, -2)
    :param max_row: 文件最大行
    :param ws: Worksheet对象
    :return: 真正写入文件的坐标，tuple格式
    """
    row, col = coord
    if col < 0:
        col = ws.max_column + col + 1
        if col < 1:
            raise ValueError(f'列号不能小于1。当前：{col}')

    if row is None:
        row = max_row + 1
    elif row < 0:
        row = max_row + row + 1
        if row < 1:
            raise ValueError(f'行号不能小于1。当前：{row}')

    return row, col


def data_to_list_or_dict_simplify(data):
    """将传入的数据转换为列表或字典形式，不添加前后列数据
    :param data: 要处理的数据
    :return: 转变成列表或字典形式的数据
    """
    if data is None:
        data = tuple()
    elif not isinstance(data, (list, tuple, dict)):
        data = (data,)
    return data


def data_to_list_or_dict(recorder, data):
    """将传入的数据转换为列表或字典形式，添加前后列数据
    :param recorder: BaseRecorder对象
    :param data: 要处理的数据
    :return: 转变成列表或字典形式的数据
    """
    if data is None:
        data = tuple()

    elif not isinstance(data, (list, tuple, dict)):
        data = (data,)

    if not (recorder.before or recorder.after):
        return data

    if isinstance(data, (list, tuple)):
        return_list = []
        for i in (recorder.before, data, recorder.after):
            if isinstance(i, dict):
                return_list.extend(list(i.values()))
            elif i is None:
                pass
            elif isinstance(i, list):
                return_list.extend(i)
            elif isinstance(i, tuple):
                return_list.extend(list(i))
            else:
                return_list.extend([str(i)])

        return return_list

    elif isinstance(data, dict):
        if not recorder.before:
            pass
        elif isinstance(recorder.before, dict):
            data = {**recorder.before, **data}
        elif isinstance(recorder.before, (list, tuple)):
            data1 = list(recorder.before)
            data1.extend(data.values())
            data = data1

        if not recorder.after:
            return data

        elif isinstance(data, dict):
            if isinstance(recorder.after, dict):
                data = {**data, **recorder.after}
            elif isinstance(recorder.after, (list, tuple)):
                data = list(data)
                data.extend(recorder.after)

        elif isinstance(data, list):
            if isinstance(recorder.after, dict):
                data.extend(recorder.after.values())
            elif isinstance(recorder.after, (list, tuple)):
                data.extend(recorder.after)

        return data


def get_and_set_csv_header(recorder, is_filler=False):
    """在写入数据时，先获取表头，如果文件不存在就新建，如果空文件且数据为dict，自动增加表头"""
    new = False
    add_header = False
    if recorder._file_exists or Path(recorder.path).exists():
        from csv import reader
        with open(recorder.path, 'r', newline='', encoding=recorder.encoding) as f:
            u = reader(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
            try:
                for _ in range(recorder._header_row):
                    header = next(u)
                if not header or not any([i for i in header]):
                    header = False  # 有表头行，但表头行是空的
            except StopIteration:  # 文件是空的
                if is_filler:
                    coord, first_data = recorder._data[0]
                    if coord != (None, True):
                        recorder._header = False
                        return
                    else:
                        first_data = first_data[0]
                else:
                    first_data = recorder._data[0]

                if isinstance(first_data, dict):
                    first_data = first_data.keys()
                    new = True
                    add_header = True
                header = ok_list(first_data)

    else:
        new = True
        if not is_filler:
            first_data = recorder._data[0]
        else:
            coord, first_data = recorder._data[0]
            if coord == (None, True):
                first_data = first_data[0]
            else:
                first_data = None

        if isinstance(first_data, dict):
            first_data = first_data.keys()
            new = True
            add_header = True

        header = ok_list(first_data) if first_data else False

    if new:
        with open(recorder.path, 'w', newline='', encoding=recorder.encoding) as f:
            if add_header:
                from csv import writer
                csv_write = writer(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
                csv_write.writerow(header)

    recorder._file_exists = True
    recorder._header = remove_list_end_Nones(header)


def get_and_set_xlsx_header(recorder, new_sheet, first_data, ws, is_filler=False):
    """获取xlsx文件的表头，如果是空的旧文件，插入表头
    :return: tuple：(开始行号, 是否新插入的表头)
    """
    if is_filler:
        coord, first_data = first_data
        other_data = first_data[1:]
        first_data = first_data[0]
        if coord != (None, True):
            coord = None
    else:
        other_data = False
        coord = True

    if (recorder._header_row == 0 and recorder._fit_header and isinstance(first_data, dict) and
            (new_sheet or not any([i.value for i in ws[1]]) and ws.max_row == 1)):
        for k, v in first_data.items():
            if isinstance(k, int):
                ws.cell(1, k).value = v
        for r, d in enumerate(other_data, start=2):
            if isinstance(d, dict):
                for k, v in d.items():
                    if isinstance(k, int):
                        ws.cell(r, k).value = v
            else:
                ws.append(ok_list(d, True))
        return 1, True

    if recorder._header.get(ws.title, None) is not None:
        return 0, False

    new_header = False
    begin = 0

    if new_sheet:
        if not coord:
            header = False
        elif isinstance(first_data, dict):
            header = ok_list(list(first_data.keys()), True)
            ws.append(header)
        else:
            header = ok_list(first_data, True)

    elif not any([i.value for i in ws[1]]) and ws.max_row == 1:  # 空旧文件
        if not coord:
            header = False
        else:
            if isinstance(first_data, dict):
                header = ok_list(first_data.keys(), True)
            else:
                begin = 1
                header = ok_list(first_data, True)
            for n, i in enumerate(header, 1):
                ws.cell(recorder._header_row, n).value = i
            for r, d in enumerate(other_data, start=2):
                if recorder._fit_header and isinstance(d, dict):
                    for k, v in d.items():
                        ind = header.index(k)
                        if ind != -1:
                            ws.cell(recorder._header_row + r, ind + 1).value = v
                        elif recorder._auto_new_col:
                            header.append(k)
                            col = ws.max_column + 1
                            ws.cell(recorder._header_row, col).value = k
                            ws.cell(recorder._header_row + r, col).value = v

                else:
                    ws.append(ok_list(d, True))
            new_header = True

    else:  # 旧有内容的sheet
        header = recorder._header.get(ws.title, None)
        if not header:
            if recorder._header_row:
                header = [i.value for i in ws[recorder._header_row]]
            else:
                header = [get_column_letter(k) for k in range(1, ws.max_row + 1)]

    if header:
        recorder._header[ws.title] = remove_list_end_Nones(header)
    else:
        recorder._header[ws.title] = False

    return begin, new_header


def handle_new_sheet(recorder, ws, data, first_wrote):
    need_set_style = (recorder._header_row,)
    if recorder._header.get(ws.title, None) and recorder._header_row > 0:
        for c, v in enumerate(recorder._header.get(ws.title, None), 1):
            ws.cell(row=recorder._header_row, column=c, value=v)

    elif (recorder._header_row > 0 and (data and (isinstance(data[0], dict) or (
            data[0] and isinstance(data[0], (list, tuple, set) and isinstance(data[0][0], dict)))))):
        # 第一个数据是dict，设置表头
        header = data[0].keys() if isinstance(data[0], dict) else data[0][0].keys()
        for c, v in enumerate(header, 1):
            ws.cell(row=recorder._header_row, column=c, value=v)
        if recorder._fit_header:
            recorder._header[ws.title] = list(header)

    elif data and recorder._header_row in (1, 0):  # 未设置表头，将第一个数据写到第一行
        first_wrote = True
        if data[0] and isinstance(data[0], (list, tuple)) and isinstance(data[0][0], (list, tuple)):
            for r, d in enumerate(data[0], 1):
                for c, v in enumerate(ok_list(d, excel=True), 1):
                    ws.cell(row=r, column=c, value=v)
            need_set_style = tuple(range(1, len(data[0]) + 1))
        else:  # 一维数据
            for c, v in enumerate(ok_list(data[0], excel=True), 1):
                ws.cell(row=1, column=c, value=v)

    if recorder._styles or recorder._col_height:
        for r in need_set_style:
            _set_style(recorder._col_height, recorder._styles, ws, r)

    return first_wrote


def _set_style(height, styles, ws, row):
    if height is not None:
        ws.row_dimensions[row].height = height

    if styles:
        if isinstance(styles, CellStyle):
            for c in ws[row]:
                styles.to_cell(c)
        else:
            for k, s in enumerate(styles, start=1):
                if s:
                    s.to_cell(ws.cell(row=row, column=k))


def get_header(recorder, table=None):
    """获取表头"""
    if recorder.type == 'xlsx':
        if table in recorder._header:
            return recorder._header[table]
        if recorder._header_row == 0 or not recorder.path or not Path(recorder.path).exists():
            return None
        wb = load_workbook(recorder.path)
        table = table or recorder.table
        if not table:
            ws = wb.active
            table = ws.title
            recorder.set.table(table)
        elif table not in wb.sheetnames:
            wb.close()
            return None
        else:
            ws = wb[table]
        header = [i.value for i in ws[recorder._header_row]]
        recorder._header[table] = False if not header else remove_list_end_Nones(header)
        wb.close()
        return recorder._header[table]

    if recorder.type == 'csv':
        if recorder._header:
            return recorder._header
        if recorder._header_row == 0 or not recorder.path or not Path(recorder.path).exists():
            return None
        from csv import reader
        with open(recorder.path, 'r', newline='', encoding=recorder.encoding) as f:
            u = reader(f, delimiter=recorder.delimiter, quotechar=recorder.quote_char)
            try:
                for _ in range(recorder._header_row):
                    header = next(u)
                if not header or not any([i for i in header]):
                    header = False  # 有表头行，但表头行是空的
            except StopIteration:  # 文件是空的
                header = None
        recorder._header = remove_list_end_Nones(header)
        return recorder._header


def create_csv(recorder):
    if not Path(recorder.path).exists():
        with open(recorder.path, 'w', newline='', encoding=recorder.encoding):
            pass
    recorder._file_exists = True


def get_wb(recorder):
    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        new_file = False
    else:
        wb = Workbook()
        new_file = True
    recorder._file_exists = True
    return wb, new_file


def get_ws(wb, table, tables, new_file):
    new_sheet = new_file
    if table is None:
        ws = wb.active
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(row=1, column=1).value:
            new_sheet = True

    elif table in tables:
        ws = wb[table]
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(row=1, column=1).value:
            new_sheet = True

    elif new_file is True:
        ws = wb.active
        tables.remove(ws.title)
        ws.title = table
        tables.append(table)
        new_sheet = True

    else:
        ws = wb.create_sheet(title=table)
        tables.append(table)
        new_sheet = True

    return ws, new_sheet


def remove_list_end_Nones(in_list):
    """去除列表后面所有None
    :param in_list: 要处理的list
    """
    h = []
    flag = True
    for i in in_list[::-1]:
        if flag:
            if i is None:
                continue
            else:
                flag = False
        h.append(i)
    return h[::-1]


def fix_openpyxl_bug(recorder, wb, ws, table):
    """尝试解决openpyxl的bug"""
    cell = ws.cell(1, 1)
    if cell.value is None:
        cell.value = ''
    wb.save(recorder.path)
    wb.close()
    wb = load_workbook(recorder.path)
    return wb, wb[table] if table else wb.active


def format_signs(signs):
    if not isinstance(signs, (list, tuple, set)):
        signs = (signs,)
    return signs


def get_tables(path):
    wb = load_workbook(path)
    tables = wb.sheetnames
    wb.close()
    return tables


class FillerDict(dict):
    pass


class FillerList(list):
    pass
