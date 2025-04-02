# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from time import sleep

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .base import BaseRecorder
from .setter import FillerSetter, get_sign_col, get_key_cols
from .style.cell_style import CellStyle, NoneStyle, CellStyleCopier
from .tools import (parse_coord, get_usable_coord, process_content, data_to_list_or_dict, ok_list, get_usable_coord_int,
                    data_to_list_or_dict_simplify, get_and_set_csv_header, get_wb, get_ws, get_and_set_xlsx_header,
                    create_csv, FillerDict, FillerList, get_header, format_signs, fix_openpyxl_bug)


class Filler(BaseRecorder):
    def __init__(self, path=None, cache_size=None,
                 key_cols=True, begin_row=2, sign_col=True, data_col=None, sign=None, deny_sign=False):  # 即将废弃
        """用于处理表格文件的工具
        :param path: 保存的文件路径
        :param cache_size: 每接收多少条记录写入文件，传入0表示不自动保存
        """
        super().__init__(path=path, cache_size=cache_size)
        self._delimiter = ','  # csv文件分隔符
        self._quote_char = '"'  # csv文件引用符

        self._follow_styles = False
        self._key_cols = True
        self._begin_row = 2
        self._sign_col = True
        self._data_col = 1
        self._sign = (None,)
        self._deny_sign = False

        self._header_row = 1
        self._fit_header = False
        self._style_data = {}

        # -------即将废弃-------
        if not data_col:
            self._data_col = sign_col if sign_col else 1
        self._key_cols = get_key_cols(key_cols)
        self._begin_row = begin_row
        self._sign_col = get_sign_col(sign_col)
        self._sign = format_signs(sign)
        self._deny_sign = deny_sign
        # ---------------------

        s = CellStyle()
        s.font.set_color("0000FF")
        s.font.set_underline('single')
        self._link_style = s

    @property
    def signs(self):
        """按这个值筛选需要的行纳入rows"""
        return self._sign

    @property
    def deny_sign(self):
        """返回是否反向匹配sign"""
        return self._deny_sign

    @property
    def key_cols(self):
        """返回作为关键字的列或列的集合"""
        return self._key_cols

    @property
    def sign_col(self):
        """返回用于判断是否已填数据的列"""
        return self._sign_col

    @property
    def data_col(self):
        """返回用于填充数据的列"""
        return self._data_col

    @property
    def begin_row(self):
        """返回数据开始的行号，用于获取rows，从1开始"""
        return self._begin_row

    @property
    def set(self):
        """返回用于设置属性的对象"""
        if self._setter is None:
            self._setter = FillerSetter(self)
        return self._setter

    @property
    def delimiter(self):
        """返回csv文件分隔符"""
        return self._delimiter

    @property
    def quote_char(self):
        """返回csv文件引用符"""
        return self._quote_char

    @property
    def header(self):
        """返回表头"""
        return get_header(self)

    def key2col(self, key, table=None):
        """接收关键字，返回该关键字在表头中的序号
        :param key: 关键字
        :param table: xlsx格式时可指定在某个sheet中获取，为None时使用对象设置的值
        :return: 序号，从1开始，没有该关键字时报错
        """
        header = get_header(self, table=table)
        if header and key in header:
            return header.index(key) + 1
        raise RuntimeError(f'没有找到指定表格或表头中没有该值：{key}')

    def add_data(self, data, coord='newline', table=None):
        """添加数据，每次添加一行数据，可指定坐标、列号或行号
        coord只输入数字（行号）时，列号为self.data_col值，如 3；
        输入列号，或没有行号的坐标时，表示新增一行，列号为此时指定的，如'c'、',3'、(None, 3)、'None,3'；
        输入 'newline' 时，表示新增一行，列号为self.data_col值；
        输入行列坐标时，填写到该坐标，如'a3'、'3,1'、(3,1)、[3,1]；
        输入的行号列号可以是负数，代表从下往上数，-1是倒数第一行，如'a-3'、(-3, -3)
        :param data: 要添加的内容，任意格式
        :param coord: 要添加数据的坐标，可输入行号、列号或行列坐标，如'a3'、7、(3, 1)、[3, 1]、'c'
        :param table: 要写入的数据表，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        while self._pause_add:  # 等待其它线程写入结束
            sleep(.1)

        if not isinstance(data, (list, tuple)):
            data = (data,)

        to = self._data
        if coord in ('cover_style', 'replace_style', 'set_width', 'set_height'):
            to = self._style_data

        elif coord not in ('set_link', 'set_img'):
            coord = parse_coord(coord, self.data_col)
            if not data:
                data = ([],)
                self._data_count += 1
            # 一维数组
            elif isinstance(data, dict) or (
                    isinstance(data, (list, tuple)) and not isinstance(data[0], (list, tuple, dict))):
                data = (data_to_list_or_dict(self, data),)
                self._data_count += 1
            else:  # 二维数组
                if self.after or self.before:
                    data = [data_to_list_or_dict(self, d) for d in data]
                else:
                    data = [data_to_list_or_dict_simplify(d) for d in data]
                self._data_count += len(data)

        else:
            self._data_count += 1

        if self._type == 'xlsx':
            if table is None:
                table = self._table
            elif isinstance(table, bool):
                table = None
            to.setdefault(table, []).append((coord, data))

        elif not self._type:
            raise RuntimeError('请设置文件路径。')

        else:
            to.append((coord, data))

        if 0 < self.cache_size <= self._data_count:
            self.record()

    def set_link(self, coord, link, content=None, table=None):
        """为单元格设置超链接
        :param coord: 单元格坐标
        :param link: 超链接，为None时删除链接
        :param content: 单元格内容
        :param table: 数据表名，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        self.add_data((coord, link, content), 'set_link', table)

    def set_style(self, coord, style, replace=True, table=None):
        """为单元格设置样式，可批量设置范围内的单元格
        :param coord: 单元格坐标，输入数字可设置整行，输入列名字符串可设置整列，输入'A1:C5'、'a:d'、'1:5'格式可设置指定范围
        :param style: CellStyle对象，为None则清除单元格样式
        :param replace: 是否直接替换已有样式，运行效率较高，但不能单独修改某个属性
        :param table: 数据表名，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        s = 'replace_style' if replace else 'cover_style'
        self.add_data((coord, style), s, table)

    def set_img(self, coord, img_path, width=None, height=None, table=None):
        """
        :param coord: 单元格坐标
        :param img_path: 图片路径
        :param width: 图片宽
        :param height: 图片高
        :param table: 数据表名，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        if isinstance(img_path, Path):
            img_path = str(img_path)
        self.add_data((coord, img_path, width, height), 'set_img', table)

    def set_row_height(self, row, height, table=None):
        """设置行高，可设置连续多行
        :param row: 行号，可传入范围，如'1:4'
        :param height: 行高
        :param table: 数据表名，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        self.add_data((row, height), 'set_height', table)

    def set_col_width(self, col, width, table=None):
        """设置列宽，可设置连续多列
        :param col: 列号，数字或字母，可传入范围，如'1:4'、'a:d'
        :param width: 列宽
        :param table: 数据表名，仅支持xlsx格式。为None表示用set.table()方法设置的值，为bool表示活动的表格
        :return: None
        """
        self.add_data((col, width), 'set_width', table)

    def rows(self, as_dict=True, key_cols=None, begin_row=None, sign_col=None,
             signs=None, deny_sign=None, count=None):
        """设置文件路径
        :param as_dict: 是否以返回dict，为False则返回list
        :param key_cols: 作为关键字的列，可以是多列
        :param begin_row: 数据开始的行，默认表头一行
        :param sign_col: 用于判断是否已填数据的列
        :param signs: 按这个值判断是否已填数据，可用list, tuple, set设置多个
        :param deny_sign: 是否反向匹配sign，即筛选指不是sign的行
        :param count: 获取多少条数据，None为所有
        :return: FillerDict或FillerList对象
        """
        if self.type == 'csv':
            return get_csv_rows(self, as_dict, key_cols=key_cols, begin_row=begin_row,
                                sign_col=sign_col, sign=signs, deny_sign=deny_sign, count=count)
        elif self.type == 'xlsx':
            return get_xlsx_rows(self, as_dict, key_cols=key_cols, begin_row=begin_row,
                                 sign_col=sign_col, sign=signs, deny_sign=deny_sign, count=count)

    def _record(self):
        """记录数据"""
        if self.type == 'xlsx':
            self._to_xlsx()
        elif self.type == 'csv':
            self._to_csv()
        self._style_data = {}

    def _to_xlsx(self):
        """填写数据到xlsx文件"""
        wb, new_file = get_wb(self)
        tables = [i.title for i in wb.worksheets]

        for table in {}.fromkeys(list(self._data.keys()) + list(self._style_data.keys())):
            ws, new_sheet = get_ws(wb, table, tables, new_file)
            if self._data:
                begin = get_and_set_xlsx_header(self, new_sheet, self._data[table][0], ws, True)[0]
            empty = not any([i.value for i in ws[1]]) and ws.max_row == 1
            header = self._header.get(ws.title, None) if self._fit_header else None

            if new_file:
                wb, ws = fix_openpyxl_bug(self, wb, ws, ws.title)
            new_file = False

            if self._data:  # 处理表头
                if not header:
                    data_method = set_data_to_ws_without_header
                elif self._follow_styles:
                    data_method = set_data_to_ws_with_header_and_style
                else:
                    data_method = set_data_to_ws_with_header_not_style

                for table_data in self._data[table][begin:]:
                    if table_data[0] == 'set_link':
                        set_link_to_ws(ws, table_data, empty, self)
                    elif table_data[0] == 'set_img':
                        set_img_to_ws(ws, table_data, empty, self)
                    else:
                        data_method(ws, table_data, empty, header, follow_styles=self._follow_styles)
                    empty = False

            if self._style_data:
                for table_data in self._style_data[table]:
                    set_style_to_ws(ws, table_data, self)

        wb.save(self.path)
        wb.close()

    def _to_csv(self):
        """填写数据到csv文件"""
        if self._header is not None and not self._file_exists:
            create_csv(self)
        elif self._header is None and self._header_row > 0:
            get_and_set_csv_header(self, True)

        with open(self.path, 'r', encoding=self.encoding) as f:
            reader = csv_reader(f, delimiter=self.delimiter, quotechar=self.quote_char)
            lines = list(reader)
            lines_count = len(lines)

            header_len = len(self._header) if self._fit_header and self._header else None
            for i in self._data:
                if i[0] == 'set_link':
                    coord = parse_coord(i[1][0], self.data_col)
                    now_data = (f'=HYPERLINK("{i[1][1]}","{i[1][2] or i[1][1]}")',)

                elif i[0] == 'set_img':
                    continue

                else:
                    coord = i[0]
                    now_data = i[1]

                row, col = get_usable_coord_int(coord, lines_count, len(lines[0]) if lines_count else 1)
                now_data = (now_data,) if not isinstance(now_data[0], (list, tuple, dict)) else now_data

                for r, data in enumerate(now_data, row):
                    for _ in range(r - lines_count):  # 若行数不够，填充行数
                        lines.append([])
                        lines_count += 1
                    row_num = r - 1

                    if isinstance(data, dict):
                        if self._fit_header and self._header:
                            # 若列数不够，填充空列
                            lines[row_num].extend([''] * (header_len - len(lines[row_num])))
                            for k, h in enumerate(self._header):
                                if h in data:
                                    lines[row_num][k] = data[h]
                            lines[row_num] = ok_list(lines[row_num])
                            continue

                        else:
                            data = ok_list(data.values())

                    # 若列数不够，填充空列
                    lines[row_num].extend([''] * (col - len(lines[row_num]) + len(data) - 1))
                    for k, j in enumerate(data):  # 填充数据
                        lines[row_num][col + k - 1] = process_content(j)

            writer = csv_writer(open(self.path, 'w', encoding=self.encoding, newline=''),
                                delimiter=self.delimiter, quotechar=self.quote_char)
            writer.writerows(lines)

    # ----------- 即将废弃 -----------
    @property
    def keys(self):
        """返回一个列表，由未执行的行数据组成。每行的格式为第一位为行号，其余为 key 列的值。
        eg.[3, '张三', 20]
        """
        return self.rows(as_dict=False)

    @property
    def dict_keys(self):
        """返回一个列表，由未执行的行数据组成。每行的格式为dict，'row' 值为行号，其余值为第一行数据。
        如第一行数据为空，则用列号为值。如果begin_row为1，用列名作为值。
        eg.{'row': 2, 'name': '张三', 'C': '男'}
        """
        return self.rows()

    @property
    def sign(self):
        return self._sign


def get_xlsx_rows(filler, as_dict, key_cols, begin_row, sign_col, sign, deny_sign, count):
    """返回key列内容，第一位为行号，其余为key列的值
    如果as_dict为True，返回dict格式，value为第一行值，值为空或begin_row为1时用列号，'row' 值为行号
    eg.[3, '名称', 'id']
    :param filler: 记录器对象
    :param as_dict: 是否以dict格式返回数据
    :param key_cols: 作为关键字的列，可以是多列，为None使用对象已有设置
    :param begin_row: 数据开始的行，默认表头一行，为None使用对象已有设置
    :param sign_col: 用于作为筛选条件的列，为None使用对象已有设置
    :param sign: 按这个值判断是否已填数据，为None使用对象已有设置
    :param deny_sign: 是否反向匹配sign，即筛选指不是sign的行，为None使用对象已有设置
    :param count: 获取多少条数据，None为所有
    :return: 关键字组成的列表或字典
    """
    key_cols = get_key_cols(key_cols) if key_cols is not None else filler.key_cols
    begin_row = begin_row if begin_row is not None else filler.begin_row
    sign_col = get_sign_col(sign_col) if sign_col is not None else filler.sign_col
    sign = format_signs(sign) if sign is not None else filler._sign
    deny_sign = deny_sign if deny_sign is not None else filler.deny_sign

    wb = load_workbook(filler.path, data_only=True, read_only=True)
    if filler.table and filler.table not in [i.title for i in wb.worksheets]:
        raise RuntimeError(f'xlsx文件未包含此工作表：{filler.table}')
    ws = wb[filler.table] if filler.table else wb.active

    if ws.max_column is None:  # 遇到过read_only时无法获取列数的文件
        wb.close()
        wb = load_workbook(filler.path, data_only=True)
        ws = wb[filler.table] if filler.table else wb.active

    rows = ws.rows
    if as_dict:
        if filler._header_row:
            header = [c.value for c in ws[filler._header_row]]
            u = next((i for i, s in enumerate(header[::-1]) if s is not None), None)
            u = - u if u else None
            header = [c if c else get_column_letter(k) for k, c in enumerate(header[:u], 1)
                      if key_cols is True or k in key_cols]
            if len(header) != len(set(header)):
                raise RuntimeError(f'表头重复项：{", ".join({t for t in header if header.count(t) > 1})}。')
        else:
            header = [get_column_letter(k) for k in range(1, ws.max_column + 1)]

        method = _make_dict_data

    else:
        header = None
        method = _make_list_data

    try:
        for _ in range(begin_row - 1):
            next(rows)
    except StopIteration:
        return []

    # ---------------------------------------------------------

    if sign_col is True or sign_col > ws.max_column:  # 获取所有行
        if count:
            rows = list(rows)[:count]
        if key_cols is True:  # 获取整行
            res = [method(ind, [i.value for i in row], header, None)
                   for ind, row in enumerate(rows, begin_row)]
        else:  # 只获取对应的列
            res = [method(ind, [row[i - 1].value for i in key_cols], header, None)
                   for ind, row in enumerate(rows, begin_row)]

    else:  # 获取符合条件的行
        if count:
            res = _handle_xlsx_rows_with_count(key_cols, deny_sign, method, header, rows,
                                               begin_row, sign_col, sign, count)
        else:
            res = _handle_xlsx_rows_without_count(key_cols, deny_sign, method, header, rows, begin_row, sign_col, sign)

    wb.close()
    return res


def get_csv_rows(filler, as_dict, key_cols, begin_row, sign_col, sign, deny_sign, count):
    """返回key列内容，第一位为行号，其余为key列的值，
    如果as_dict为True，返回dict格式，value为第一行值，值为空或begin_row为1时用列号，'row'值为行号
    eg.[3, '名称', 'id']
    :param filler: 记录器对象
    :param as_dict: 是否以dict格式返回数据
    :param key_cols: 作为关键字的列，可以是多列，为None使用对象已有设置
    :param begin_row: 数据开始的行，默认表头一行，为None使用对象已有设置
    :param sign_col: 用于作为筛选条件的列，为None使用对象已有设置
    :param sign: 按这个值判断是否已填数据，为None使用对象已有设置
    :param deny_sign: 是否反向匹配sign，即筛选指不是sign的行，为None使用对象已有设置
    :param count: 获取多少条数据，None为所有
    :return: 关键字组成的列表或字典
    """
    key_cols = get_key_cols(key_cols) if key_cols is not None else filler.key_cols
    begin_row = begin_row if begin_row is not None else filler.begin_row
    sign_col = get_sign_col(sign_col) if sign_col is not None else filler.sign_col
    deny_sign = deny_sign if deny_sign is not None else filler.deny_sign
    sign = format_signs(sign) if sign is not None else filler._sign
    sign = ['' if i is None else str(i) for i in sign]

    begin_row -= 1
    res = []

    with open(filler.path, 'r', encoding=filler.encoding) as f:
        reader = csv_reader(f, delimiter=filler.delimiter, quotechar=filler.quote_char)
        lines = list(reader)
        if not lines:
            return res

        if as_dict:
            if filler._header_row:
                header = [x if x else get_column_letter(k) for k, x in enumerate(lines[filler._header_row - 1], 1)
                          if key_cols is True or k in key_cols]
                if len(header) != len(set(header)):
                    raise RuntimeError(f'表头重复项：{",".join({t for t in header if header.count(t) > 1})}。')
                method = _make_dict_data
            else:
                header = None
                method = _make_zero_header_dict_data

        else:
            header = None
            method = _make_list_data

        if sign_col is True:  # 获取所有行
            for ind, line in enumerate(lines[begin_row:count + 1 if count else None], begin_row + 1):
                if key_cols is True:  # 获取整行
                    res.append(method(ind, line, header, ''))
                else:  # 只获取对应的列
                    res.append(method(ind, [line[i - 1] for i in key_cols], header, ''))

        else:  # 获取符合条件的行
            sign_col -= 1
            if count:
                _handle_csv_rows_with_count(lines, begin_row, sign_col, sign, deny_sign,
                                            key_cols, res, method, header, count)
            else:
                _handle_csv_rows_without_count(lines, begin_row, sign_col, sign, deny_sign,
                                               key_cols, res, method, header)

    return res


def set_link_to_ws(ws, data, empty, filler):
    max_row = 0 if empty else ws.max_row
    coord = parse_coord(data[1][0], filler.data_col)
    row, col = get_usable_coord(coord, max_row, ws)
    cell = ws.cell(row, col)
    has_link = True if cell.hyperlink else Filler
    cell.hyperlink = None if data[1][1] is None else process_content(data[1][1], True)
    if data[1][2] is not None:
        cell.value = process_content(data[1][2], True)
    if data[1][1]:
        if filler._link_style:
            filler._link_style.to_cell(cell, replace=False)
    elif has_link:
        NoneStyle().to_cell(cell, replace=False)


def set_img_to_ws(ws, data, empty, filler):
    max_row = 0 if empty else ws.max_row
    coord, img_path, width, height = data[1]
    coord = parse_coord(coord, filler.data_col)
    row, col = get_usable_coord(coord, max_row, ws)

    from openpyxl.drawing.image import Image
    img = Image(img_path)
    if width and height:
        img.width = width
        img.height = height
    elif width:
        img.height = int(img.height * (width / img.width))
        img.width = width
    elif height:
        img.width = int(img.width * (height / img.height))
        img.height = height
    col = get_column_letter(col)
    ws.add_image(img, f'{col}{row}')


def set_data_to_ws_with_header_not_style(ws, data, empty, header, follow_styles):
    max_row = 0 if empty else ws.max_row
    row, col = get_usable_coord(data[0], max_row, ws)
    now_data = (data[1],) if not isinstance(data[1][0], (list, tuple, dict)) else data[1]

    for r, i in enumerate(now_data, row):
        if isinstance(i, dict):
            for k, h in enumerate(header, 1):
                if h in i:
                    ws.cell(r, k).value = process_content(i[h], True)

        else:
            for key, j in enumerate(i):
                ws.cell(r, col + key).value = process_content(j, True)


def set_data_to_ws_with_header_and_style(ws, data, empty, header, follow_styles):
    max_row = 0 if empty else ws.max_row
    row, col = get_usable_coord(data[0], max_row, ws)
    now_data = (data[1],) if not isinstance(data[1][0], (list, tuple, dict)) else data[1]

    if data[0][0]:  # 非新行
        styles = []
        for r, i in enumerate(now_data, row):
            if isinstance(i, dict):
                style = []
                for k, h in enumerate(header, 1):
                    if h in i:
                        ws.cell(r, k).value = process_content(i[h], True)
                        style.append(k)
                styles.append(style)
            else:
                for key, j in enumerate(i):
                    ws.cell(r, col + key).value = process_content(j, True)
                styles.append(range(1, len(i) + 1))

        if max_row >= row - 1:
            copy_some_row_style(ws, row, styles)

    else:  # 新行
        for r, i in enumerate(now_data, row):
            if isinstance(i, dict):
                for k, h in enumerate(header, 1):
                    if h in i:
                        ws.cell(r, k).value = process_content(i[h], True)
            else:
                for key, j in enumerate(i):
                    ws.cell(r, col + key).value = process_content(j, True)

        if max_row >= row - 1:
            copy_full_row_style(ws, row, now_data)


def set_data_to_ws_without_header(ws, data, empty, header, follow_styles):
    max_row = 0 if empty else ws.max_row
    row, col = get_usable_coord(data[0], max_row, ws)
    now_data = (data[1],) if not isinstance(data[1][0], (list, tuple, dict)) else data[1]

    for r, i in enumerate(now_data, row):
        if isinstance(i, dict):
            i = i.values()
        for key, j in enumerate(i):
            ws.cell(r, col + key).value = process_content(j, True)

    if follow_styles and max_row >= row - 1:
        copy_part_row_style(ws, row, now_data, col) if data[0][0] else copy_full_row_style(ws, row, now_data)


def copy_some_row_style(ws, row, styles):
    """复制上一行指定列样式到后续行中"""
    _row_styles = [CellStyleCopier(i) for i in ws[row - 1]]
    for r, i in enumerate(styles, row):
        for c in i:
            if _row_styles[c - 1]:
                _row_styles[c - 1].to_cell(ws.cell(row=r, column=c))


def copy_full_row_style(ws, row, now_data):
    """复制上一行整行样式到新行中"""
    _row_styles = [CellStyleCopier(i) for i in ws[row - 1]]
    height = ws.row_dimensions[row - 1].height
    for r, i in enumerate(now_data, row):
        for k, s in enumerate(_row_styles, start=1):
            if s:
                s.to_cell(ws.cell(row=r, column=k))
        ws.row_dimensions[r].height = height


def copy_part_row_style(ws, row, now_data, col):
    """复制上一行局部（连续）样式到后续行中"""
    _row_styles = [CellStyleCopier(i) for i in ws[row - 1]]
    for r, i in enumerate(now_data, row):
        for c in range(len(i)):
            if _row_styles[c + col - 1]:
                _row_styles[c + col - 1].to_cell(ws.cell(row=r, column=c + col))


def set_style_to_ws(ws, data, filler):
    """批量设置单元格格式到sheet"""
    if data[0] in ('replace_style', 'cover_style'):
        mode = data[0] == 'replace_style'
        coord = data[1][0]
        style = NoneStyle() if data[1][1] is None else data[1][1]
        if isinstance(coord, int) or (isinstance(coord, str) and coord.isdigit()):
            for c in ws[coord]:
                style.to_cell(c, replace=mode)
            return

        elif isinstance(coord, str):
            if ':' in coord:
                for c in ws[coord]:
                    for cc in c:
                        style.to_cell(cc, replace=mode)
                return
            elif coord.isdigit() or coord.isalpha():
                for c in ws[coord]:
                    style.to_cell(c, replace=mode)
                return

        coord = parse_coord(coord, filler.data_col)
        row, col = get_usable_coord(coord, ws.max_row, ws)
        style.to_cell(ws.cell(row, col), replace=mode)

    elif data[0] == 'set_width':
        col, width = data[1]
        if isinstance(col, int):
            col = get_column_letter(col)
        for c in col.split(':'):
            if c.isdigit():
                c = get_column_letter(int(c))
            ws.column_dimensions[c].width = width

    elif data[0] == 'set_height':
        row, height = data[1]
        if isinstance(row, int):
            ws.row_dimensions[row].height = height
        elif isinstance(row, str):
            for r in row.split(':'):
                ws.row_dimensions[int(r)].height = height


def _handle_xlsx_rows_with_count(key_cols, deny_sign, method, header, rows, begin_row, sign_col, sign, count):
    got = 0
    res = []
    if key_cols is True:  # 获取整行
        if deny_sign:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value not in sign:
                    res.append(method(ind, [i.value for i in row], header, None))
                    got += 1
        else:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value in sign:
                    res.append(method(ind, [i.value for i in row], header, None))
                    got += 1

    else:  # 只获取对应的列
        if deny_sign:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value not in sign:
                    res.append(method(ind, [row[i - 1].value for i in key_cols], header, None))
                    got += 1
        else:
            for ind, row in enumerate(rows, begin_row):
                if got == count:
                    break
                if row[sign_col - 1].value in sign:
                    res.append(method(ind, [row[i - 1].value for i in key_cols], header, None))
                    got += 1
    return res


def _handle_xlsx_rows_without_count(key_cols, deny_sign, method, header, rows, begin_row, sign_col, sign):
    if key_cols is True:  # 获取整行
        if deny_sign:
            return [method(ind, [i.value for i in row], header, None)
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value not in sign]
        else:
            return [method(ind, [i.value for i in row], header, None)
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value in sign]

    else:  # 只获取对应的列
        if deny_sign:
            return [method(ind, [row[i - 1].value for i in key_cols], header, None)
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value not in sign]
        else:
            return [method(ind, [row[i - 1].value for i in key_cols], header, None)
                    for ind, row in enumerate(rows, begin_row)
                    if row[sign_col - 1].value in sign]


def _handle_csv_rows_with_count(lines, begin_row, sign_col, sign, deny_sign, key_cols, res, method, header, count):
    got = 0
    for ind, line in enumerate(lines[begin_row:], begin_row + 1):
        if got == count:
            break
        row_sign = '' if sign_col > len(line) - 1 else line[sign_col]
        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                res.append(method(ind, line, header, ''))
            else:  # 只获取对应的列
                res.append(method(ind, [line[i - 1] for i in key_cols], header, ''))
            got += 1


def _handle_csv_rows_without_count(lines, begin_row, sign_col, sign, deny_sign, key_cols, res, method, header):
    for ind, line in enumerate(lines[begin_row:], begin_row + 1):
        row_sign = '' if sign_col > len(line) - 1 else line[sign_col]
        if (row_sign not in sign) if deny_sign else (row_sign in sign):
            if key_cols is True:  # 获取整行
                res.append(method(ind, line, header, ''))
            else:  # 只获取对应的列
                res.append(method(ind, [line[i - 1] for i in key_cols], header, ''))


def _make_list_data(ind, data, header=None, None_val=None):
    r = FillerList(data)
    r.row = ind
    return r


def _make_dict_data(ind, data, header, None_val):
    data += [None_val] * (len(header) - len(data))
    r = FillerDict(zip(header, data))
    r.row = ind
    return r


def _make_zero_header_dict_data(ind, data, header, None_val):
    header = [get_column_letter(k) for k in range(1, len(data) + 1)]
    r = FillerDict(zip(header, data))
    r.row = ind
    return r
